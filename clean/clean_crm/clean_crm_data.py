"""CRM订单/退款清洗脚本"""
import os
import sys
import json
import re
from datetime import datetime
from collections import defaultdict

# PyInstaller打包后__file__指向临时目录
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.join(os.path.dirname(sys.executable), 'clean_crm')
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'output'))

# 导入公共工具
sys.path.insert(0, os.path.normpath(os.path.join(SCRIPT_DIR, '..')))
from utils import (
    read_file, clean_strings, remove_empty_rows,
    _parse_yyyymm, _normalize_value, write_output, _not_empty,
    save_cache, get_paths_interactive
)
from openpyxl import load_workbook


def load_crm_config():
    with open(os.path.join(SCRIPT_DIR, 'crm_config.json'), 'r', encoding='utf-8') as f:
        return json.load(f)


def load_merchant_mapping(xlsx_path):
    """从crm商户映射.xlsx加载商户ID→店铺/主体/渠道/CRM店铺"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    headers = [str(h) if h is not None else '' for h in rows[0]]
    result = {}
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        mid = str(row_dict.get('商户ID', '')).strip()
        if mid:
            result[mid] = {k: row_dict.get(k, '') for k in ['店铺', '主体', '渠道', 'CRM店铺']}
    return result


def load_mst_mapping(mst_path):
    """加载Mst映射表，返回(精确dict, 商品名dict)。
    精确dict: {商品名+单价 normalized拼接: {IP, 类目, 归属, 结算表引用}}  — 订单用
    商品名dict: {商品名: {IP, 类目, 归属, 结算表引用}}  — 退款用（退款金额≠原始单价）
    """
    wb = load_workbook(mst_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}, {}

    headers = [str(h) if h is not None else '' for h in rows[0]]
    get_fields = ['IP', '类目', '归属', '结算表引用', '引用关联（IP核对）', '引用关联（管报）']
    exact_dict = {}
    name_dict = {}
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        name = _normalize_value(row_dict.get('商品名'))
        price = _normalize_value(row_dict.get('单价'))
        entry = {f: row_dict.get(f, '') for f in get_fields}
        has_data = any(_not_empty(entry.get(f)) for f in get_fields)
        # 精确映射(商品名+单价)
        ref_key = name + price
        if ref_key.strip() and (ref_key not in exact_dict or has_data):
            exact_dict[ref_key] = entry
        # 商品名映射(仅商品名，首条有效数据优先)
        name_key = name.strip()
        if name_key and name_key not in name_dict and has_data:
            name_dict[name_key] = entry
    return exact_dict, name_dict


def find_mst_file(config):
    """从脚本目录查找Mst文件"""
    mst_name = config['Mst路径']
    local = os.path.join(SCRIPT_DIR, mst_name)
    if os.path.exists(local):
        return local
    return None


def scan_crm_files(source_dir):
    """扫描目录，识别CRM订单和退款文件"""
    result = {'订单': [], '退款': []}
    files = [f for f in os.listdir(source_dir)
             if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')]
    for fname in sorted(files):
        full_path = os.path.join(source_dir, fname)
        if 'CRM订单' in fname:
            result['订单'].append(full_path)
        elif 'CRM退款' in fname:
            result['退款'].append(full_path)
    return result


def match_mst(data, mst_dict, name_field, amount_field):
    """用商品名+金额匹配Mst映射，返回(data, unmatched_list)。未匹配统一用'商品名'字段名。"""
    get_fields = ['IP', '类目', '归属', '结算表引用', '引用关联（IP核对）', '引用关联（管报）']
    unmatched = []
    for row in data:
        name = _normalize_value(row.get(name_field))
        amount = _normalize_value(row.get(amount_field))
        ref_key = name + amount
        matched = mst_dict.get(ref_key, {})
        for f in get_fields:
            val = matched.get(f, '')
            row[f] = '' if val is None else val
        if ref_key.strip() and (not matched or not any(_not_empty(matched.get(f)) for f in get_fields)):
            unmatched.append({'商品名': row.get(name_field, ''), '金额': row.get(amount_field, '')})
    return data, unmatched


def match_mst_by_name(data, name_dict, name_field):
    """仅用商品名匹配Mst映射（退款用，因为退款金额≠原始单价），返回(data, unmatched_list)"""
    get_fields = ['IP', '类目', '归属', '结算表引用', '引用关联（IP核对）', '引用关联（管报）']
    unmatched = []
    for row in data:
        name = _normalize_value(row.get(name_field)).strip()
        matched = name_dict.get(name, {})
        for f in get_fields:
            val = matched.get(f, '')
            row[f] = '' if val is None else val
        if name and (not matched or not any(_not_empty(matched.get(f)) for f in get_fields)):
            unmatched.append({'商品名': row.get(name_field, '')})
    return data, unmatched


def _parse_time(t):
    """解析时间值为datetime，失败返回None"""
    if t is None:
        return None
    if isinstance(t, datetime):
        return t
    s = str(t).strip()
    if not s:
        return None
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                 '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def load_live_record_map(xlsx_path):
    """从真直播时间映射.xlsx加载真直播订单ID集合"""
    if not xlsx_path or not os.path.exists(xlsx_path):
        return set()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return set()
    headers = [str(h) if h is not None else '' for h in rows[0]]
    result = set()
    for row in rows[1:]:
        d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if '真直播' not in str(d.get('直播形式', '')):
            continue
        oid = str(d.get('订单ID', '')).strip()
        if oid.endswith('.0'):
            oid = oid[:-2]
        if oid:
            result.add(oid)
    return result


def determine_live_format(ref_value, time_value, live_config, live_record_set=None, order_id=None):
    """判断直播形式：
    - 结算表引用不是关键词之一 → ''（不判断）
    - 月份<202602：订单ID在映射表 →（真直播），否则→（伪直播）
    - 月份>=202602：时间落入规则区间 →（真直播），否则→（伪直播）
    返回 (直播形式str, is_new_hit bool)"""
    keywords = live_config['关键词']
    ref_str = str(ref_value).strip() if ref_value else ''
    # 结算表引用必须精确等于某个关键词
    if ref_str not in keywords:
        return '', False

    # 订单ID标准化
    oid = ''
    if order_id is not None:
        oid = str(order_id).strip()
        if oid.endswith('.0'):
            oid = oid[:-2]

    # 解析时间和月份
    dt = _parse_time(time_value)
    month = dt.strftime('%Y%m') if dt else ''

    if month and month < '202602':
        # <202602：仅用订单ID匹配映射表
        if live_record_set and oid and oid in live_record_set:
            return '（真直播）', False
        return '（伪直播）', False
    else:
        # >=202602（或无法解析月份）：用时间区间规则
        if dt is not None:
            for period in live_config['真直播时间段']:
                start = datetime.strptime(period['开始'], '%Y-%m-%d %H:%M')
                end = datetime.strptime(period['结束'], '%Y-%m-%d %H:%M')
                if start <= dt <= end:
                    return '（真直播）', True  # 新命中，追加到映射表
        return '（伪直播）', False


def append_live_records(xlsx_path, new_records):
    """将真直播记录追加到真直播时间映射.xlsx（去重by订单ID）"""
    if not new_records:
        return
    HEADERS = ['订单ID', '商品名称', '创建时间', '直播形式', '结算表引用']

    existing_ids = set()
    if xlsx_path and os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # 读取已有订单ID用于去重
        for r in range(2, ws.max_row + 1):
            oid = str(ws.cell(r, 1).value or '').strip()
            if oid.endswith('.0'):
                oid = oid[:-2]
            if oid:
                existing_ids.add(oid)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for ci, h in enumerate(HEADERS, 1):
            ws.cell(1, ci, h)

    added = 0
    for rec in new_records:
        oid = str(rec.get('订单ID', '')).strip()
        if oid.endswith('.0'):
            oid = oid[:-2]
        if not oid or oid in existing_ids:
            continue
        existing_ids.add(oid)
        row_idx = ws.max_row + 1
        for ci, h in enumerate(HEADERS, 1):
            ws.cell(row_idx, ci, rec.get(h, ''))
        added += 1

    if added > 0:
        wb.save(xlsx_path)
        print(f'  追加真直播记录: {added}条 → {os.path.basename(xlsx_path)}')
    wb.close()


def map_merchant(data, merchant_map):
    """商户ID → 店铺/主体/渠道/CRM店铺"""
    for row in data:
        mid = str(row.get('商户ID', '')).strip()
        if mid.endswith('.0'):
            mid = mid[:-2]
        info = merchant_map.get(mid, {})
        row['店铺'] = info.get('店铺', '')
        row['主体'] = info.get('主体', '')
        row['渠道'] = info.get('渠道', '')
        row['CRM店铺'] = info.get('CRM店铺', '')
    return data


def calc_adjusted_amount(row, month_field='下单月份', rules=None):
    """变动金额：按配置规则对指定IP从指定月份起乘以系数，其他=原始金额"""
    ip = str(row.get('IP', '')).strip()
    month = str(row.get(month_field, '')).strip()
    amount = row.get('金额', 0)
    try:
        amount = float(amount) if amount else 0
    except (ValueError, TypeError):
        amount = 0
    for rule in (rules or []):
        if ip == rule['IP'] and month >= rule['起始月份']:
            return round(amount * rule['系数'], 2)
    return round(amount, 2)


def dedup_rows(data):
    """完整行去重"""
    seen = set()
    result = []
    for row in data:
        key = tuple(sorted((k, str(v) if v is not None else '') for k, v in row.items()))
        if key not in seen:
            seen.add(key)
            result.append(row)
    return result


def process_crm_orders(files, mst_dict, config, merchant_map, live_record_set=None):
    """处理CRM订单，返回(输出数据list, 未匹配list, 新真直播记录list)"""
    source_fields = ['订单ID', '商户订单号', '商品名', '金额', '利息', '创建时间', '订单归属人', '商户ID']
    output_fields = ['渠道', 'CRM店铺', '店铺', '主体', '订单ID', '商户订单号', '商品名', '金额', '利息',
                     '下单月份', '订单归属人', '商户ID', 'IP', '类目', '归属', '直播形式',
                     '结算表引用', '引用关联（IP核对）', '引用关联（管报）', '变动金额']

    all_data = []
    all_unmatched = []
    new_live_records = []

    for fpath in files:
        fname = os.path.basename(fpath)
        print(f'  处理: {fname}')
        data = read_file(fpath, {'格式': 'excel'})
        print(f'    读取: {len(data)}行')
        data = clean_strings(data)
        data = remove_empty_rows(data)

        # 提取字段
        data = [{f: row.get(f, '') for f in source_fields} for row in data]

        # Mst映射
        data, unmatched = match_mst(data, mst_dict, '商品名', '金额')
        all_unmatched.extend(unmatched)

        # 直播形式 + 引用关联（IP核对）追加直播形式
        for row in data:
            fmt, is_new = determine_live_format(
                row.get('结算表引用'), row.get('创建时间'), config['直播形式'],
                live_record_set, order_id=row.get('订单ID'))
            row['直播形式'] = fmt
            if fmt:
                row['引用关联（IP核对）'] = str(row.get('引用关联（IP核对）', '') or '') + fmt
            if is_new:
                oid = str(row.get('订单ID', '')).strip()
                if oid.endswith('.0'):
                    oid = oid[:-2]
                new_live_records.append({
                    '订单ID': oid,
                    '商品名称': row.get('商品名', ''),
                    '创建时间': row.get('创建时间', ''),
                    '直播形式': '（真直播）',
                    '结算表引用': row.get('结算表引用', '')
                })
                if live_record_set is not None and oid:
                    live_record_set.add(oid)

        # 商户ID→店铺/主体
        data = map_merchant(data, merchant_map)

        # 创建时间→下单月份
        for row in data:
            row['下单月份'] = _parse_yyyymm(row.get('创建时间'))

        all_data.extend(data)

    all_data = dedup_rows(all_data)
    # 计算变动金额
    adj_rules = config.get('变动金额规则', [])
    for row in all_data:
        row['变动金额'] = calc_adjusted_amount(row, '下单月份', adj_rules)
    result = [{f: row.get(f, '') for f in output_fields} for row in all_data]
    print(f'  CRM订单合计: {len(result)}行（去重后）')
    return result, all_unmatched, new_live_records


def process_crm_refunds(files, mst_name_dict, config, merchant_map, live_record_set=None):
    """处理CRM退款，返回(输出数据list, 未匹配list)"""
    source_fields = ['订单ID', '商品名称', '金额', '支付时间', '审核时间', '状态', '商户ID']
    output_fields = ['渠道', 'CRM店铺', '店铺', '主体', '订单ID', '商品名称', '金额', '下单月份', '审核月份',
                     '状态', '商户ID', 'IP', '类目', '归属', '直播形式',
                     '结算表引用', '引用关联（IP核对）', '引用关联（管报）', '变动金额']

    all_data = []
    all_unmatched = []

    for fpath in files:
        fname = os.path.basename(fpath)
        print(f'  处理: {fname}')
        data = read_file(fpath, {'格式': 'excel'})
        print(f'    读取: {len(data)}行')
        data = clean_strings(data)
        data = remove_empty_rows(data)

        # 提取字段
        data = [{f: row.get(f, '') for f in source_fields} for row in data]

        # 过滤：状态含退款成功
        data = [r for r in data if '退款成功' in str(r.get('状态', ''))]
        print(f'    过滤后: {len(data)}行')

        # Mst映射（退款用仅商品名匹配，因为退款金额≠原始单价）
        data, unmatched = match_mst_by_name(data, mst_name_dict, '商品名称')
        all_unmatched.extend(unmatched)

        # 直播形式（退款用支付时间，按订单ID匹配）+ 引用关联追加直播形式
        for row in data:
            fmt, _ = determine_live_format(
                row.get('结算表引用'), row.get('支付时间'), config['直播形式'],
                live_record_set, order_id=row.get('订单ID'))
            row['直播形式'] = fmt
            if fmt:
                row['引用关联（IP核对）'] = str(row.get('引用关联（IP核对）', '') or '') + fmt

        # 商户ID→店铺/主体
        data = map_merchant(data, merchant_map)

        # 支付时间→下单月份，审核时间→审核月份
        for row in data:
            row['下单月份'] = _parse_yyyymm(row.get('支付时间'))
            row['审核月份'] = _parse_yyyymm(row.get('审核时间'))

        all_data.extend(data)

    all_data = dedup_rows(all_data)
    # 计算变动金额（售后按下单月份）
    adj_rules = config.get('变动金额规则', [])
    for row in all_data:
        row['变动金额'] = calc_adjusted_amount(row, '下单月份', adj_rules)
    result = [{f: row.get(f, '') for f in output_fields} for row in all_data]
    print(f'  CRM退款合计: {len(result)}行（去重后）')
    return result, all_unmatched


def pivot_crm_orders(data):
    """CRM订单汇总：按(下单月份, 渠道, CRM店铺, IP, 引用关联（IP核对）)汇总"""
    group_fields = ['下单月份', '渠道', 'CRM店铺', 'IP', '引用关联（IP核对）']
    groups = defaultdict(lambda: {'金额': 0, '利息': 0, '变动金额': 0, '订单数量': 0})
    for row in data:
        key = tuple(str(row.get(f, '')) for f in group_fields)
        amount = row.get('金额', 0)
        interest = row.get('利息', 0)
        adjusted = row.get('变动金额', 0)
        try:
            amount = float(amount) if amount else 0
        except (ValueError, TypeError):
            amount = 0
        try:
            interest = float(interest) if interest else 0
        except (ValueError, TypeError):
            interest = 0
        try:
            adjusted = float(adjusted) if adjusted else 0
        except (ValueError, TypeError):
            adjusted = 0
        groups[key]['金额'] += amount
        groups[key]['利息'] += interest
        groups[key]['变动金额'] += adjusted
        groups[key]['订单数量'] += 1

    result = []
    for key, vals in sorted(groups.items()):
        row = {group_fields[i]: key[i] for i in range(len(group_fields))}
        row['订单数量'] = vals['订单数量']
        row['金额'] = round(vals['金额'], 2)
        row['利息'] = round(vals['利息'], 2)
        row['变动金额'] = round(vals['变动金额'], 2)
        result.append(row)
    # 合计行
    total = {'下单月份': '合计', '渠道': '', 'CRM店铺': '', 'IP': '', '引用关联（IP核对）': '',
             '订单数量': sum(r['订单数量'] for r in result),
             '金额': round(sum(r['金额'] for r in result), 2),
             '利息': round(sum(r['利息'] for r in result), 2),
             '变动金额': round(sum(r['变动金额'] for r in result), 2)}
    result.append(total)
    return result


def pivot_crm_refunds(data):
    """CRM退款汇总：按(下单月份, 审核月份, 渠道, CRM店铺, IP, 引用关联（IP核对）)汇总"""
    group_fields = ['下单月份', '审核月份', '渠道', 'CRM店铺', 'IP', '引用关联（IP核对）']
    groups = defaultdict(lambda: {'金额': 0, '变动金额': 0, '退款数量': 0})
    for row in data:
        key = tuple(str(row.get(f, '')) for f in group_fields)
        amount = row.get('金额', 0)
        adjusted = row.get('变动金额', 0)
        try:
            amount = float(amount) if amount else 0
        except (ValueError, TypeError):
            amount = 0
        try:
            adjusted = float(adjusted) if adjusted else 0
        except (ValueError, TypeError):
            adjusted = 0
        groups[key]['金额'] += amount
        groups[key]['变动金额'] += adjusted
        groups[key]['退款数量'] += 1

    result = []
    for key, vals in sorted(groups.items()):
        row = {group_fields[i]: key[i] for i in range(len(group_fields))}
        row['退款数量'] = vals['退款数量']
        row['金额'] = round(vals['金额'], 2)
        row['变动金额'] = round(vals['变动金额'], 2)
        result.append(row)
    # 合计行
    total = {'下单月份': '合计', '审核月份': '', '渠道': '', 'CRM店铺': '', 'IP': '', '引用关联（IP核对）': '',
             '退款数量': sum(r['退款数量'] for r in result),
             '金额': round(sum(r['金额'] for r in result), 2),
             '变动金额': round(sum(r['变动金额'] for r in result), 2)}
    result.append(total)
    return result


def pivot_crm_combined(orders, refunds):
    """合并汇总：订单按下单月份、退款按审核月份，汇总订单金额/利息/退款金额/净金额"""
    group_fields = ['月份', '渠道', 'CRM店铺', 'IP', '引用关联（IP核对）']
    groups = defaultdict(lambda: {'订单金额': 0, '利息': 0, '退款金额': 0, '订单收入': 0, '售后退款': 0})

    for row in orders:
        key = (str(row.get('下单月份', '')), str(row.get('渠道', '')),
               str(row.get('CRM店铺', '')), str(row.get('IP', '')),
               str(row.get('引用关联（IP核对）', '')))
        try:
            groups[key]['订单金额'] += float(row.get('金额', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['利息'] += float(row.get('利息', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['订单收入'] += float(row.get('变动金额', 0) or 0)
        except (ValueError, TypeError):
            pass

    for row in refunds:
        key = (str(row.get('审核月份', '')), str(row.get('渠道', '')),
               str(row.get('CRM店铺', '')), str(row.get('IP', '')),
               str(row.get('引用关联（IP核对）', '')))
        try:
            groups[key]['退款金额'] += float(row.get('金额', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['售后退款'] += float(row.get('变动金额', 0) or 0)
        except (ValueError, TypeError):
            pass

    result = []
    for key, vals in sorted(groups.items()):
        row = {group_fields[i]: key[i] for i in range(len(group_fields))}
        row['订单金额'] = round(vals['订单金额'], 2)
        row['利息'] = round(vals['利息'], 2)
        row['退款金额'] = round(vals['退款金额'], 2)
        row['净金额'] = round(vals['订单金额'] - vals['退款金额'], 2)
        row['订单收入'] = round(vals['订单收入'], 2)
        row['售后退款'] = round(vals['售后退款'], 2)
        result.append(row)
    # 合计行
    total = {'月份': '合计', '渠道': '', 'CRM店铺': '', 'IP': '', '引用关联（IP核对）': '',
             '订单金额': round(sum(r['订单金额'] for r in result), 2),
             '利息': round(sum(r['利息'] for r in result), 2),
             '退款金额': round(sum(r['退款金额'] for r in result), 2),
             '净金额': round(sum(r['净金额'] for r in result), 2),
             '订单收入': round(sum(r['订单收入'] for r in result), 2),
             '售后退款': round(sum(r['售后退款'] for r in result), 2)}
    result.append(total)
    return result


def pivot_crm_ip_summary(orders, refunds):
    """按(月份, 渠道, CRM店铺, IP)汇总。订单用下单月份，退款用审核月份（当月退款）。
    订单金额/退款金额=原始金额，订单收入/售后退款=变动金额（含0.85系数）。"""
    groups = defaultdict(lambda: {'订单金额': 0, '利息': 0, '退款金额': 0, '订单收入': 0, '售后退款': 0})
    for row in orders:
        key = (str(row.get('下单月份', '') or ''),
               str(row.get('渠道', '') or ''),
               str(row.get('CRM店铺', '') or ''),
               str(row.get('IP', '') or ''))
        try:
            groups[key]['订单金额'] += float(row.get('金额', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['利息'] += float(row.get('利息', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['订单收入'] += float(row.get('变动金额', 0) or 0)
        except (ValueError, TypeError):
            pass
    for row in refunds:
        key = (str(row.get('审核月份', '') or ''),
               str(row.get('渠道', '') or ''),
               str(row.get('CRM店铺', '') or ''),
               str(row.get('IP', '') or ''))
        try:
            groups[key]['退款金额'] += float(row.get('金额', 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            groups[key]['售后退款'] += float(row.get('变动金额', 0) or 0)
        except (ValueError, TypeError):
            pass
    result = []
    for key, agg in sorted(groups.items()):
        result.append({
            '月份': key[0], '渠道': key[1], '店铺': key[2], 'IP': key[3],
            '订单金额': round(agg['订单金额'], 2),
            '利息': round(agg['利息'], 2),
            '退款金额': round(agg['退款金额'], 2),
            '净金额': round(agg['订单金额'] - agg['退款金额'], 2),
            '订单收入': round(agg['订单收入'], 2),
            '售后退款': round(agg['售后退款'], 2),
        })
    return result


def _resolve_source_dirs(raw_paths):
    """解析数据源路径：含YYYYMM店铺子目录时自动展开"""
    source_dirs = []
    for p in raw_paths:
        p = p.strip().strip('"').strip("'").rstrip('\\')
        if not p or not os.path.isdir(p):
            if p:
                print(f'  路径跳过(不存在): {p}')
            continue
        subs = sorted([
            os.path.join(p, d) for d in os.listdir(p)
            if os.path.isdir(os.path.join(p, d)) and re.match(r'^\d{6}店铺$', d)
        ])
        if subs:
            source_dirs.extend(subs)
            print(f'  自动展开: {os.path.basename(p)} → {", ".join(os.path.basename(s) for s in subs)}')
        else:
            source_dirs.append(p)
    return source_dirs


def main(input_paths=None):
    config = load_crm_config()

    if input_paths:
        raw_paths = input_paths
        for p in raw_paths:
            print(f'  输入路径: {p}')
        save_cache(raw_paths)
    elif len(sys.argv) > 1:
        print(f'\n拖放模式')
        raw_paths = sys.argv[1:]
        for p in raw_paths:
            print(f'  输入路径: {p}')
        save_cache(raw_paths)
    else:
        raw_paths = get_paths_interactive()
        if not raw_paths:
            return
        save_cache(raw_paths)

    source_dirs = _resolve_source_dirs(raw_paths)
    if not source_dirs:
        print('无有效路径')
        return

    # 查找Mst文件
    mst_path = find_mst_file(config)
    if not mst_path:
        print(f'找不到Mst映射文件: {config["Mst路径"]}')
        return
    print(f'  Mst映射文件: {mst_path}')

    mst_dict, mst_name_dict = load_mst_mapping(mst_path)
    print(f'  Mst映射条目: {len(mst_dict)}(精确), {len(mst_name_dict)}(商品名)')

    # 加载商户映射
    merchant_path = os.path.join(SCRIPT_DIR, config['商户映射路径'])
    merchant_map = load_merchant_mapping(merchant_path)
    print(f'  商户映射: {len(merchant_map)}条（来自{config["商户映射路径"]}）')

    # 加载真直播时间映射
    live_map_path = os.path.join(SCRIPT_DIR, config.get('真直播时间映射路径', ''))
    live_record_set = load_live_record_map(live_map_path)
    if live_record_set:
        print(f'  真直播映射记录: {len(live_record_set)}条（来自{config.get("真直播时间映射路径", "")}）')

    # 无规则提醒
    has_time_rules = bool(config.get('直播形式', {}).get('真直播时间段', []))
    if not has_time_rules and not live_record_set:
        ans = input('⚠ 当前无真直播时间规则且无历史映射记录，含关键词的记录将全标为伪直播，继续？(Y/n) ').strip().lower()
        if ans == 'n':
            return

    all_orders = []
    all_refunds = []
    all_unmatched = []
    all_new_live_records = []

    for source_dir in source_dirs:
        print(f'\n>>> 数据源: {os.path.basename(source_dir)}')
        crm_files = scan_crm_files(source_dir)
        print(f'  CRM订单文件: {len(crm_files["订单"])}个')
        print(f'  CRM退款文件: {len(crm_files["退款"])}个')

        if crm_files['订单']:
            orders, unmatched, new_live_recs = process_crm_orders(crm_files['订单'], mst_dict, config, merchant_map, live_record_set)
            all_orders.extend(orders)
            all_unmatched.extend(unmatched)
            all_new_live_records.extend(new_live_recs)

        if crm_files['退款']:
            refunds, unmatched = process_crm_refunds(crm_files['退款'], mst_name_dict, config, merchant_map, live_record_set)
            all_refunds.extend(refunds)
            all_unmatched.extend(unmatched)

    # 追加新真直播记录到映射表
    if all_new_live_records:
        append_live_records(live_map_path, all_new_live_records)

    # 输出
    output_dir = OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, '清洗_私域订单.xlsx')

    sheets = {}
    if all_orders:
        sheets['CRM订单'] = all_orders
        order_summary = pivot_crm_orders(all_orders)
        sheets['订单汇总'] = order_summary
        print(f'  订单汇总: {len(order_summary)}行')
    if all_refunds:
        sheets['CRM退款'] = all_refunds
        refund_summary = pivot_crm_refunds(all_refunds)
        sheets['退款汇总'] = refund_summary
        print(f'  退款汇总: {len(refund_summary)}行')

    if all_orders or all_refunds:
        combined = pivot_crm_combined(all_orders, all_refunds)
        sheets['合并汇总'] = combined
        print(f'  合并汇总: {len(combined)}行')
        ip_summary = pivot_crm_ip_summary(all_orders, all_refunds)
        sheets['IP汇总'] = ip_summary
        print(f'  IP汇总: {len(ip_summary)}行')

    if all_unmatched:
        seen = set()
        unique = []
        for u in all_unmatched:
            key = tuple(sorted(u.items()))
            if key not in seen:
                seen.add(key)
                unique.append(u)
        sheets['未匹配商品'] = unique
        print(f'\n未匹配商品: {len(unique)}种（去重后）')

    write_output(sheets, output_path)

    print('\n' + '=' * 50)
    print('CRM清洗完成!')
    print(f'  CRM订单: {len(all_orders)}行')
    print(f'  CRM退款: {len(all_refunds)}行')
    print(f'  输出: {output_path}')
    print('=' * 50)


if __name__ == '__main__':
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    print('=' * 50)
    print('  CRM数据清洗工具')
    print(f'  工作目录: {SCRIPT_DIR}')
    print('=' * 50)
    try:
        main()
    except Exception as e:
        print(f'\n*** 运行出错: {e} ***')
        import traceback
        traceback.print_exc()
    input('\n按回车键退出...')
