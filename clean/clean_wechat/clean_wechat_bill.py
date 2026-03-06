"""微信账单清洗模块"""
import os
import sys
import csv
import calendar
from collections import defaultdict
from openpyxl import load_workbook

from utils import _to_number, _parse_date, write_output, read_output_sheets

# === 常量 ===

# 脚本/exe所在目录
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.join(os.path.dirname(sys.executable), 'clean_wechat')
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'output'))

# 商户映射文件路径
MERCHANT_XLSX = os.path.join(SCRIPT_DIR, 'wechat商户映射.xlsx')

# 业务类型 → 明细
BIZ_TYPE_MAP = {
    "交易": "微信收入",
    "扣除交易手续费": "微信手续费",
    "退款": "微信退款",
    "提现": "提现",
    "网银充值": "提现",
}

# 提现类业务类型
WITHDRAW_TYPES = {"提现", "网银充值"}

# 输出列顺序
WITHDRAW_COLUMNS = ["月份", "店铺", "日期", "明细", "IP", "渠道", "账号名称", "日收入", "日支出"]
NON_WITHDRAW_COLUMNS = ["月份", "日期", "明细", "IP", "渠道", "账号名称", "日收入", "日支出"]


def _load_merchant_map():
    """从xlsx读取商户映射，返回 {商户ID: (店铺, 账号名称)}"""
    if not os.path.exists(MERCHANT_XLSX):
        print(f'*** 商户映射文件不存在: {MERCHANT_XLSX} ***')
        return {}
    wb = load_workbook(MERCHANT_XLSX, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            mid = str(row[0]).strip()
            shop = str(row[1]).strip() if row[1] else mid
            account = str(row[2]).strip() if row[2] else mid
            result[mid] = (shop, account)
    wb.close()
    return result


def scan_files(input_path, merchant_ids):
    """扫描目录，找含"现金"的子目录中匹配商户ID+基本账户的CSV文件。
    返回 list of (file_path, merchant_id)"""
    input_path = input_path.strip().strip('"').strip("'").rstrip('\\').rstrip('/')
    if not os.path.isdir(input_path):
        print(f"路径不存在: {input_path}")
        return []

    # 确定搜索目录：目录名含"现金"→直接搜索；否则搜子目录
    search_dirs = []
    if "现金" in os.path.basename(input_path):
        search_dirs.append(input_path)
    else:
        for name in os.listdir(input_path):
            sub = os.path.join(input_path, name)
            if os.path.isdir(sub) and "现金" in name:
                search_dirs.append(sub)

    if not search_dirs:
        print(f"未找到含'现金'的目录: {input_path}")
        return []

    results = []
    for d in search_dirs:
        for fname in sorted(os.listdir(d)):
            if not fname.lower().endswith('.csv'):
                continue
            if "基本账户" not in fname:
                continue
            for mid in merchant_ids:
                if mid in fname:
                    results.append((os.path.join(d, fname), mid))
                    break
    return results


def read_and_clean(file_path):
    """读取微信CSV（utf-8-sig），去掉字段key和value前导反引号。返回 list of dict"""
    with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            cleaned = {}
            for k, v in row.items():
                clean_key = k.lstrip('`').strip()
                clean_val = v.lstrip('`').strip() if v else ''
                cleaned[clean_key] = clean_val
            rows.append(cleaned)
    return rows


def _month_last_day(yyyymm):
    """给定yyyymm字符串，返回该月最后一天 yyyy-mm-dd"""
    if not yyyymm or len(yyyymm) < 6:
        return ''
    try:
        year = int(yyyymm[:4])
        month = int(yyyymm[4:6])
        _, last_day = calendar.monthrange(year, month)
        return f'{year:04d}-{month:02d}-{last_day:02d}'
    except (ValueError, TypeError):
        return ''


def process_wechat(files, merchant_map):
    """处理所有文件数据。
    返回 (withdraw_rows, non_withdraw_rows)
    """
    all_rows = []
    for file_path, mid in files:
        fname = os.path.basename(file_path)
        data = read_and_clean(file_path)
        print(f"  读取: {fname} → {len(data)}行")

        shop, account = merchant_map.get(mid, (mid, mid))

        for row in data:
            # 提取时间（跳过无法解析的行，如CSV末尾汇总行）
            月份, 日期 = _parse_date(row.get('记账时间', ''))
            if not 月份:
                continue

            # 业务类型映射
            biz_type = row.get('业务类型', '').strip()
            明细 = BIZ_TYPE_MAP.get(biz_type, biz_type)

            # 收支金额解析
            收支类型 = row.get('收支类型', '').strip()
            amount = _to_number(row.get('收支金额(元)', '')) or 0.0
            日收入 = amount if 收支类型 == '收入' else 0.0
            日支出 = amount if 收支类型 == '支出' else 0.0
            # 提现类如在收入列，改为支出负数
            if biz_type in WITHDRAW_TYPES and 收支类型 == '收入':
                日收入 = 0.0
                日支出 = -amount

            all_rows.append({
                '月份': 月份, '店铺': shop, '日期': 日期,
                '明细': 明细, 'IP': '其他', '渠道': '微信',
                '账号名称': account, '日收入': 日收入, '日支出': 日支出,
                '_业务类型': biz_type,
            })

    # 分为提现和非提现
    withdraw = [r for r in all_rows if r['_业务类型'] in WITHDRAW_TYPES]
    non_withdraw = [r for r in all_rows if r['_业务类型'] not in WITHDRAW_TYPES]
    for r in withdraw:
        del r['_业务类型']
    for r in non_withdraw:
        del r['_业务类型']

    return withdraw, non_withdraw


def aggregate_withdraw(rows):
    """提现汇总（按日）"""
    groups = defaultdict(lambda: {'日收入': 0.0, '日支出': 0.0})
    keys = ['月份', '店铺', '日期', '明细', 'IP', '渠道', '账号名称']
    for row in rows:
        key = tuple(row[k] for k in keys)
        groups[key]['日收入'] += row['日收入']
        groups[key]['日支出'] += row['日支出']

    result = []
    for key, agg in groups.items():
        r = dict(zip(keys, key))
        r['日收入'] = round(agg['日收入'], 2)
        r['日支出'] = round(agg['日支出'], 2)
        result.append(r)
    result.sort(key=lambda x: (x['月份'], x['店铺'], x['日期']))
    return result


def aggregate_non_withdraw(rows):
    """不含提现汇总（按月，日期=月末）"""
    groups = defaultdict(lambda: {'日收入': 0.0, '日支出': 0.0})
    keys = ['月份', '明细', 'IP', '渠道', '账号名称']
    for row in rows:
        key = tuple(row[k] for k in keys)
        groups[key]['日收入'] += row['日收入']
        groups[key]['日支出'] += row['日支出']

    result = []
    for key, agg in groups.items():
        r = dict(zip(keys, key))
        r['日期'] = _month_last_day(r['月份'])
        r['日收入'] = round(agg['日收入'], 2)
        r['日支出'] = round(agg['日支出'], 2)
        result.append(r)
    result.sort(key=lambda x: (x['月份'], x['账号名称'], x['明细']))
    return result


def _reorder(rows, columns):
    """按指定列顺序重新排列dict"""
    return [{col: r.get(col, '') for col in columns} for r in rows]


def main(input_paths):
    """主入口，接收路径列表（与shop/crm统一接口）"""
    # 加载商户映射
    merchant_map = _load_merchant_map()
    if not merchant_map:
        print('商户映射为空，无法继续')
        return
    print(f'商户映射: {len(merchant_map)}个')

    # 取第一个路径
    input_path = input_paths[0] if input_paths else ''
    if not input_path:
        print('未提供路径')
        return

    # 扫描文件
    print(f'\n扫描目录: {input_path}')
    files = scan_files(input_path, list(merchant_map.keys()))
    if not files:
        print('未找到匹配的CSV文件（需包含商户ID且包含"基本账户"）')
        return
    print(f'找到 {len(files)} 个文件')

    # 处理
    withdraw_rows, non_withdraw_rows = process_wechat(files, merchant_map)
    print(f'提现记录: {len(withdraw_rows)}条, 非提现记录: {len(non_withdraw_rows)}条')

    # 汇总
    withdraw_agg = aggregate_withdraw(withdraw_rows)
    non_withdraw_agg = aggregate_non_withdraw(non_withdraw_rows)
    print(f'提现汇总: {len(withdraw_agg)}行, 不含提现汇总: {len(non_withdraw_agg)}行')

    # 提取批次月份
    batch_months = set()
    for r in withdraw_agg + non_withdraw_agg:
        if r.get('月份'):
            batch_months.add(r['月份'])
    if batch_months:
        print(f'当前批次月份: {", ".join(sorted(batch_months))}')

    # 输出路径（固定文件名，支持累积存档）
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, '微信账单清洗.xlsx')

    # 存档合并：加载已有数据，按批次月份去重后合并
    existing = read_output_sheets(output_path)
    old_withdraw = existing.get('提现汇总', [])
    old_non_withdraw = existing.get('不含提现汇总', [])

    # 旧数据修正（xlsx读回None/int→''/float）
    for rows in (old_withdraw, old_non_withdraw):
        for r in rows:
            for k in r:
                if r[k] is None:
                    r[k] = ''
            r['日收入'] = float(r['日收入'] or 0)
            r['日支出'] = float(r['日支出'] or 0)

    if batch_months and (old_withdraw or old_non_withdraw):
        old_w_count, old_nw_count = len(old_withdraw), len(old_non_withdraw)
        old_withdraw = [r for r in old_withdraw if r.get('月份', '') not in batch_months]
        old_non_withdraw = [r for r in old_non_withdraw if r.get('月份', '') not in batch_months]
        print(f'  存档: 旧提现{old_w_count}→{len(old_withdraw)}行, 旧非提现{old_nw_count}→{len(old_non_withdraw)}行')

    # 合并 旧(去重后) + 新
    all_withdraw = old_withdraw + _reorder(withdraw_agg, WITHDRAW_COLUMNS)
    all_non_withdraw = old_non_withdraw + _reorder(non_withdraw_agg, NON_WITHDRAW_COLUMNS)
    all_withdraw.sort(key=lambda x: (str(x.get('月份', '')), str(x.get('店铺', '')), str(x.get('日期', ''))))
    all_non_withdraw.sort(key=lambda x: (str(x.get('月份', '')), str(x.get('账号名称', '')), str(x.get('明细', ''))))

    sheets = {
        '提现汇总': all_withdraw,
        '不含提现汇总': all_non_withdraw,
    }
    write_output(sheets, output_path)

    # 统计（全量）
    w_inc = sum(float(r.get('日收入') or 0) for r in all_withdraw)
    w_exp = sum(float(r.get('日支出') or 0) for r in all_withdraw)
    nw_inc = sum(float(r.get('日收入') or 0) for r in all_non_withdraw)
    nw_exp = sum(float(r.get('日支出') or 0) for r in all_non_withdraw)
    print(f'\n处理完成!')
    print(f'  提现汇总(全量): {len(all_withdraw)}行, 收入={w_inc:.2f}, 支出={w_exp:.2f}')
    print(f'  非提现汇总(全量): {len(all_non_withdraw)}行, 收入={nw_inc:.2f}, 支出={nw_exp:.2f}')
    print(f'  输出: {output_path}')
