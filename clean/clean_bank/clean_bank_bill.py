import os, sys, json, re
from collections import defaultdict

if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.join(os.path.dirname(sys.executable), 'clean_bank')
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'output'))
sys.path.insert(0, os.path.normpath(os.path.join(SCRIPT_DIR, '..')))
from utils import write_output, _parse_date, _to_number


CONFIG_PATH = os.path.join(SCRIPT_DIR, 'bank_config.json')


def _load_config():
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def _load_mst_mapping(mst_path):
    """加载 Mst_银行账单映射.xlsx → list[{名称, 摘要, 明细, IP, 渠道}]"""
    from openpyxl import load_workbook
    wb = load_workbook(mst_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_data = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip() if c is not None else '' for c in row]
        if headers is None:
            headers = vals
            continue
        if len(vals) >= 5:
            rows_data.append({
                '名称': vals[0],
                '摘要': vals[1],
                '明细': vals[2],
                'IP': vals[3],
                '渠道': vals[4]
            })
    wb.close()
    return rows_data


def _match_mst(name, summary, mst_data):
    """匹配银行映射表：精确→模糊→仅名称→未匹配"""
    name = str(name).strip()
    summary = str(summary).strip()

    # 按名称筛选候选行
    candidates = [r for r in mst_data if r['名称'] == name]
    if not candidates:
        return '', '', ''

    # 1.精确匹配
    for c in candidates:
        if c['摘要'] == summary:
            return c['明细'], c['IP'], c['渠道']

    # 2.模糊匹配：银行摘要包含映射摘要，或映射摘要包含银行摘要
    for c in candidates:
        if c['摘要'] and (c['摘要'] in summary or summary in c['摘要']):
            return c['明细'], c['IP'], c['渠道']

    # 3.仅名称匹配：该名称下所有行的明细/IP相同
    details = set(c['明细'] for c in candidates if c['明细'])
    ips = set(c['IP'] for c in candidates if c['IP'])
    channels = set(c['渠道'] for c in candidates if c['渠道'])
    if len(details) == 1 and len(ips) == 1:
        return details.pop(), ips.pop(), channels.pop() if len(channels) == 1 else ''

    # 4.未匹配
    return '', '', ''


def _match_store(filename, config):
    """文件名中的公司名→店铺"""
    for keyword, store in config['store_mapping'].items():
        if keyword in filename:
            return store
    return '银行_未知'


def _read_bank_excel(filepath):
    """读取银行xlsx：跳过元数据行，自动寻找列头行"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 寻找列头行：同时包含"交易日"和"借方金额"的行
    header_idx = None
    for i, row in enumerate(all_rows):
        vals = [str(c).strip() if c is not None else '' for c in row]
        if '交易日' in vals and '借方金额' in vals:
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = [str(h).strip() if h is not None else '' for h in all_rows[header_idx]]
    data = []
    for row in all_rows[header_idx + 1:]:
        d = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                d[headers[j]] = val
        if any(v is not None and str(v).strip() for v in row):
            data.append(d)
    return data


def _process(files, config, mst_data):
    """处理所有银行文件"""
    all_rows = []
    for filepath in files:
        fname = os.path.basename(filepath)
        store = _match_store(fname, config)
        print(f"  处理: {fname} → {store}")

        raw = _read_bank_excel(filepath)
        if not raw:
            print(f"    跳过（无数据）")
            continue

        for row in raw:
            # 日期
            date_val = row.get('交易日', '')
            yyyymm, date_str = _parse_date(date_val)
            row['月份'] = yyyymm
            row['日期'] = date_str
            row['店铺'] = store

            # 收入/支出：借方=支出，贷方=收入（标准银行记账）
            debit = _to_number(row.get('借方金额', '')) or 0
            credit = _to_number(row.get('贷方金额', '')) or 0
            # 如果没有独立的借方/贷方列，尝试用交易金额
            if debit == 0 and credit == 0:
                amt = _to_number(row.get('交易金额', '')) or 0
                if amt > 0:
                    credit = amt
                elif amt < 0:
                    debit = abs(amt)
            row['收入'] = credit
            row['支出'] = debit
            row['借方金额'] = debit if debit else ''
            row['贷方金额'] = credit if credit else ''

            # Mst映射
            payee = str(row.get('收(付)方名称', row.get('收（付）方名称', ''))).strip()
            summary = str(row.get('摘要', row.get('交易摘要', ''))).strip()
            detail, ip, channel = _match_mst(payee, summary, mst_data)
            row['明细'] = detail
            row['IP'] = ip
            row['渠道'] = channel if channel else config['channel_name_default']
            row['收(付)方名称'] = payee

            # 摘要关键词规则（覆盖Mst结果）
            for rule in config.get('summary_rules', []):
                if rule['摘要包含'] in summary:
                    row['明细'] = rule['明细']
                    row['IP'] = rule.get('IP', row['IP'])
                    break

            # 明细重命名（如 资金互转→提现）
            rename = config.get('detail_rename', {})
            if row['明细'] in rename:
                row['明细'] = rename[row['明细']]

            # 明细→IP默认值
            ip_defaults = config.get('detail_ip_defaults', {})
            if row['明细'] in ip_defaults and not row['IP']:
                row['IP'] = ip_defaults[row['明细']]

            # 明细+摘要→渠道
            ch_rules = config.get('detail_channel_rules', {})
            if row['明细'] in ch_rules:
                for rule in ch_rules[row['明细']]:
                    if rule['摘要包含'] in summary:
                        row['渠道'] = rule['渠道']
                        break

            # 日期格式统一为 yyyy-mm-dd
            row['交易日'] = date_str

        all_rows.extend(raw)
    return all_rows


def _build_fund_summary(rows, config):
    """资金汇总（银行逐行输出）"""
    cols = config['fund_summary_columns']
    result = []
    for r in rows:
        row = {}
        for c in cols:
            if c == '日期':
                row[c] = r.get('日期', '')
            else:
                row[c] = r.get(c, '')
        result.append(row)
    result.sort(key=lambda r: (r.get('店铺', ''), r.get('日期', '')))
    return result


def main(input_paths):
    config = _load_config()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 寻找映射表和账单文件
    mst_path = None
    files = []
    for p in input_paths:
        if not os.path.isdir(p):
            continue
        for f in os.listdir(p):
            fp = os.path.join(p, f)
            if not os.path.isfile(fp) or f.startswith('~$'):
                continue
            if f == config['mst_file']:
                mst_path = fp
            elif config['file_pattern'] in f and f.endswith(('.xlsx', '.xls')):
                files.append(fp)

    if not mst_path:
        print("错误：未找到 Mst_银行账单映射.xlsx")
        return
    if not files:
        print("未找到银行账单文件")
        return

    print(f"加载映射表: {mst_path}")
    mst_data = _load_mst_mapping(mst_path)
    print(f"  映射条目: {len(mst_data)}")
    print(f"找到 {len(files)} 个银行账单文件")

    rows = _process(files, config, mst_data)
    if not rows:
        print("无数据可处理")
        return
    print(f"共 {len(rows)} 条记录")

    # 源数据
    extract_cols = config['extract_columns']
    extra_cols = config['source_extra_columns']
    all_cols = extract_cols + extra_cols
    source_data = [{k: r.get(k, '') for k in all_cols} for r in rows]

    # 资金汇总
    fund_summary = _build_fund_summary(rows, config)

    sheets = {'源数据': source_data, '资金汇总': fund_summary}
    output_path = os.path.join(OUTPUT_DIR, '银行账单清洗.xlsx')
    write_output(sheets, output_path)
    print(f"\n输出: {output_path}")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        paths = sys.argv[1:]
    else:
        paths = [input("请输入银行账单目录路径: ").strip().strip('"')]
    main(input_paths=paths)
