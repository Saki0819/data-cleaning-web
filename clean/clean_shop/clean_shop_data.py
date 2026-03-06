"""多渠道订单清洗脚本 — 配置驱动"""
import os
import sys
import json
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook

# PyInstaller打包后__file__指向临时目录，需用exe所在目录
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.join(os.path.dirname(sys.executable), 'clean_shop')
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'output'))

# 导入公共工具
sys.path.insert(0, os.path.normpath(os.path.join(SCRIPT_DIR, '..')))
from utils import (
    read_file, clean_strings, remove_empty_rows, _parse_yyyymm, _parse_date,
    _normalize_value, _not_empty, _to_number, write_output,
    save_cache, get_paths_interactive, _save_with_retry, read_output_sheets
)

def load_channel_rules():
    with open(os.path.join(SCRIPT_DIR, 'channel_rules.json'), 'r', encoding='utf-8') as f:
        return json.load(f)

def load_store_mapping():
    with open(os.path.join(SCRIPT_DIR, 'store_mapping.json'), 'r', encoding='utf-8') as f:
        return json.load(f)


# === Task 2: 文件扫描 ===

def scan_files(source_dir, channel_name, file_rules):
    """扫描目录，按文件识别规则分类订单/售后文件"""
    result = {'订单': [], '售后': []}
    files = [f for f in os.listdir(source_dir)
             if f.endswith(('.xlsx', '.xls', '.csv')) and not f.startswith('~$')]
    for fname in sorted(files):
        full_path = os.path.join(source_dir, fname)
        # 文件名含"退款"统一归为售后
        if '退款' in fname:
            # 确认属于该渠道（订单规则的"含"关键词匹配）
            订单含 = file_rules['订单']['含']
            订单含 = 订单含 if isinstance(订单含, list) else [订单含]
            if all(kw in fname for kw in 订单含):
                result['售后'].append(full_path)
            continue
        for dtype in ['售后', '订单']:  # 先匹配售后（更严格）
            rule = file_rules[dtype]
            含 = rule['含'] if isinstance(rule['含'], list) else [rule['含']]
            if not all(kw in fname for kw in 含):
                continue
            不含 = rule.get('不含', None)
            if 不含:
                不含 = 不含 if isinstance(不含, list) else [不含]
                if any(kw in fname for kw in 不含):
                    continue
            result[dtype].append(full_path)
            break
    return result


# === Task 4: 过滤 ===

def apply_filter(data, rule):
    """按过滤规则处理数据"""
    if '剔除字段' in rule:
        field = rule['剔除字段']
        if '剔除值' in rule:
            exclude_vals = set(rule['剔除值'])
            return [r for r in data if str(r.get(field, '')).strip() not in exclude_vals]
        elif '剔除值含' in rule:
            excludes = rule['剔除值含']
            return [r for r in data if not any(e in str(r.get(field, '')) for e in excludes)]
        elif rule.get('剔除条件') == '为空':
            return [r for r in data if _not_empty(r.get(field))]
    elif '保留字段' in rule:
        field = rule['保留字段']
        keep_vals = rule['保留值含']
        return [r for r in data if any(kv in str(r.get(field, '')) for kv in keep_vals)]
    return data


# === Task 5: 字段映射 + 月份提取 ===

def map_fields(data, field_mapping):
    """按字段映射重命名列，只保留映射中定义的字段"""
    result = []
    for row in data:
        new_row = {}
        for src, dst in field_mapping.items():
            if src in row:
                new_row[dst] = row[src]
        result.append(new_row)
    return result

def extract_month(data, month_config):
    """从来源字段提取yyyymm格式月份到目标字段"""
    src = month_config['来源字段']
    dst = month_config['目标字段']
    for row in data:
        val = row.get(src)
        row[dst] = _parse_yyyymm(val)
    return data


# === Task 6: 映射库加载 + 商品匹配 ===

def load_mapping_dict(xlsx_path, mapping_config):
    """从映射库xlsx加载指定sheet为dict {引用键: {字段: 值}}"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[mapping_config['sheet名']]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = mapping_config.get('header行', 0)
    if header_row >= len(rows):
        return {}
    headers = [str(h) if h is not None else '' for h in rows[header_row]]

    key_cols = mapping_config['引用键']
    get_cols = mapping_config['获取字段']

    result = {}
    for row in rows[header_row + 1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        key_parts = [_normalize_value(row_dict.get(k)) for k in key_cols]
        ref_key = ''.join(key_parts)
        if not ref_key.strip():
            continue
        new_entry = {f: row_dict.get(f, '') for f in get_cols}
        # 不用空值行覆盖已有非空数据（防止append_unmatched写入的占位行覆盖有效行）
        if ref_key in result and not any(_not_empty(new_entry.get(f)) for f in get_cols):
            continue
        result[ref_key] = new_entry
    return result


def match_products(data, mapping_dict, key_fields, get_fields):
    """用key_fields拼接引用键，从mapping_dict匹配get_fields。返回(data, unmatched_list)。
    匹配到但IP/类目为空的也视为未匹配（映射库待补充）。"""
    unmatched = []
    for row in data:
        key_parts = [_normalize_value(row.get(k)) for k in key_fields]
        ref_key = ''.join(key_parts)
        matched = mapping_dict.get(ref_key, {})
        for f in get_fields:
            val = matched.get(f, '')
            row[f] = '' if val is None else val
        if ref_key.strip():
            if not matched or not any(_not_empty(matched.get(f)) for f in get_fields):
                unmatched.append({'引用键': ref_key, **{k: row.get(k, '') for k in key_fields}})
    return data, unmatched


def build_name_mapping(xlsx_path, mapping_config):
    """从映射库按商品名（第一个引用键）构建 {商品名: {IP, 类目, 订单区分}} 索引，
    用于售后在订单编号匹配失败时按商品名兜底。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[mapping_config['sheet名']]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = mapping_config.get('header行', 0)
    if header_row >= len(rows):
        return {}
    headers = [str(h) if h is not None else '' for h in rows[header_row]]

    name_col = mapping_config['引用键'][0]  # 第一个引用键=商品名列
    get_cols = mapping_config['获取字段']

    result = {}
    for row in rows[header_row + 1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        name_val = row_dict.get(name_col)
        if name_val is None:
            continue
        name = str(name_val).strip()
        if not name or name in result:
            continue
        entry = {f: (row_dict.get(f) if row_dict.get(f) is not None else '') for f in get_cols}
        if any(_not_empty(entry.get(f)) for f in get_cols):
            result[name] = entry
    return result


# === Task 7: 售后补充 ===

def _normalize_oid(val):
    """标准化订单编号：去除.0后缀（Excel读取数值型ID常见问题）"""
    s = str(val).strip() if val else ''
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _best_order_match(aftersale_row, candidates):
    """从候选订单中选最佳匹配：优先商品名+IP都匹配的，再IP匹配的，最后兜底。"""
    name = str(aftersale_row.get('商品名', '')).strip()
    # 1. 商品名匹配 + 有IP
    if name:
        for c in candidates:
            if str(c.get('商品名', '')).strip() == name and _not_empty(c.get('IP')):
                return c
    # 2. 任意有IP的
    for c in candidates:
        if _not_empty(c.get('IP')):
            return c
    # 3. 商品名匹配（无IP）
    if name:
        for c in candidates:
            if str(c.get('商品名', '')).strip() == name:
                return c
    # 4. 兜底
    return candidates[0]


def supplement_aftersale(aftersales, orders, global_month_map=None):
    """售后通过订单编号从订单匹配IP/类目/订单区分/下单月份，已有值不覆盖。
    同订单号多商品时优先匹配商品名相同的订单行。
    下单月份缺失时额外从global_month_map（订单月份映射）兜底查找。"""
    order_map = {}
    for o in orders:
        oid = _normalize_oid(o.get('订单编号', ''))
        if oid:
            order_map.setdefault(oid, []).append(o)

    supplement_fields = ['IP', '类目', '订单区分', '下单月份']
    for row in aftersales:
        oid = _normalize_oid(row.get('订单编号', ''))
        candidates = order_map.get(oid, [])
        if candidates:
            matched = _best_order_match(row, candidates)
            for f in supplement_fields:
                if not _not_empty(row.get(f)):
                    row[f] = matched.get(f, '')
        # 下单月份仍空时从订单月份映射兜底
        if not _not_empty(row.get('下单月份')) and global_month_map and oid:
            row['下单月份'] = global_month_map.get(oid, '')
        # 最终兜底：从售后自身的订单支付时间提取
        if not _not_empty(row.get('下单月份')):
            pay_time = row.get('订单支付时间', '')
            if pay_time:
                ym, _ = _parse_date(pay_time)
                if ym:
                    row['下单月份'] = ym
    return aftersales


def load_global_order_months(global_xlsx_path):
    """加载订单月份映射。优先从本地缓存文件读取，不存在则从全域数据提取并保存。"""
    cache_path = os.path.join(SCRIPT_DIR, '订单月份映射.xlsx')
    if os.path.exists(cache_path):
        result = _load_month_cache(cache_path)
        print(f'  订单月份映射(本地): {len(result)}条')
        return result
    # 首次从全域数据提取
    if not os.path.exists(global_xlsx_path):
        print(f'  全域数据文件不存在: {global_xlsx_path}')
        return {}
    result, by_channel = _extract_months_from_global(global_xlsx_path)
    _save_month_cache(cache_path, by_channel)
    print(f'  订单月份映射(全域提取): {len(result)}条, 已保存到本地')
    return result


def _load_month_cache(cache_path):
    """从本地缓存xlsx读取所有sheet的 订单编号→下单月份"""
    wb = load_workbook(cache_path, read_only=True, data_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:
                result[_normalize_oid(row[0])] = str(row[1]).strip()
    wb.close()
    return result


def _extract_months_from_global(global_xlsx_path):
    """从全域数据xlsx提取订单编号→下单月份，返回 (flat_dict, by_channel_dict)"""
    wb = load_workbook(global_xlsx_path, read_only=True, data_only=True)
    flat = {}
    by_channel = {}  # {渠道: [(订单编号, 下单月份), ...]}
    for sn in wb.sheetnames:
        if not sn.startswith('Details_'):
            continue
        # 从sheet名提取渠道: "Details_百度订单" → "百度"
        ch_name = sn.replace('Details_', '').replace('订单', '').replace('售后', '')
        ws = wb[sn]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row1:
            continue
        headers = [str(h) if h else '' for h in row1]
        oid_idx = None
        for name in ['订单编号', '主订单编号', '订单号']:
            if name in headers:
                oid_idx = headers.index(name)
                break
        month_idx = headers.index('下单月份') if '下单月份' in headers else None
        if oid_idx is None or month_idx is None:
            continue
        by_channel.setdefault(ch_name, [])
        for row in ws.iter_rows(min_row=2, values_only=True):
            oid = row[oid_idx] if oid_idx < len(row) else None
            month = row[month_idx] if month_idx < len(row) else None
            if oid and month:
                oid_s = _normalize_oid(oid)
                month_s = str(month).strip()
                flat[oid_s] = month_s
                by_channel[ch_name].append((oid_s, month_s))
    wb.close()
    return flat, by_channel


def _save_month_cache(cache_path, by_channel):
    """保存订单月份映射到本地xlsx，按渠道分sheet"""
    wb = Workbook()
    first = True
    for ch_name, pairs in sorted(by_channel.items()):
        if not pairs:
            continue
        if first:
            ws = wb.active
            ws.title = ch_name
            first = False
        else:
            ws = wb.create_sheet(ch_name)
        ws.append(['订单编号', '下单月份'])
        seen = set()
        for oid, month in pairs:
            if oid not in seen:
                seen.add(oid)
                ws.append([oid, month])
    if first:
        ws = wb.active
        ws.title = '空'
    wb.save(cache_path)


def update_order_month_cache(all_orders):
    """将本次处理的订单数据追加到本地月份映射缓存"""
    cache_path = os.path.join(SCRIPT_DIR, '订单月份映射.xlsx')
    if not all_orders:
        return
    # 按渠道分组本次新增
    new_by_channel = {}
    for row in all_orders:
        ch = str(row.get('渠道', ''))
        oid = _normalize_oid(row.get('订单编号', ''))
        month = str(row.get('下单月份', '')).strip()
        if oid and month:
            new_by_channel.setdefault(ch, []).append((oid, month))
    if not new_by_channel:
        return

    if os.path.exists(cache_path):
        wb = load_workbook(cache_path)
    else:
        wb = Workbook()
        wb.active.title = '临时'

    added_total = 0
    for ch_name, pairs in new_by_channel.items():
        if ch_name in wb.sheetnames:
            ws = wb[ch_name]
        else:
            ws = wb.create_sheet(ch_name)
            ws.append(['订单编号', '下单月份'])
        # 收集已有的订单编号
        existing = set()
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val:
                existing.add(_normalize_oid(val))
        added = 0
        for oid, month in pairs:
            if oid not in existing:
                existing.add(oid)
                ws.append([oid, month])
                added += 1
        added_total += added

    # 删除临时sheet
    if '临时' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['临时']

    if added_total:
        _save_with_retry(wb, cache_path)
        print(f'  订单月份映射已更新: +{added_total}条')
    else:
        wb.close()


# === Task 8: 店铺名 + 透视汇总 ===

def extract_store_info(filename, channel_key, store_mapping):
    """从文件名提取渠道显示名和店铺名。
    文件名格式: '渠道 [店铺名] [密码后缀] [售后相关后缀] [时间戳].ext'
    """
    渠道覆盖 = store_mapping.get('渠道名覆盖', {})
    店铺覆盖 = store_mapping.get('店铺名覆盖', {})
    channel_display = 渠道覆盖.get(channel_key, channel_key)

    name = os.path.splitext(filename)[0]
    # 去除密码后缀（纯字母数字≥5位）
    name = re.sub(r'\s+[A-Za-z0-9]{5,}$', '', name)
    # 去除售后/退款相关后缀和时间戳
    name = re.sub(r'\s*[-_]?退款.*$', '', name)
    name = re.sub(r'\s*导?出?售后单?.*?[-\d_ ]*$', '', name)
    parts = name.split()
    store_name = parts[-1] if len(parts) > 1 else name

    for key, val in 店铺覆盖.items():
        if key in store_name:
            store_name = val
            break

    full_name = f'{channel_display}_{store_name}'
    # 最终映射：按完整店铺名覆盖（解决同名在不同渠道需不同结果）
    最终映射 = store_mapping.get('店铺名最终映射', {})
    full_name = 最终映射.get(full_name, full_name)
    return (channel_display, full_name)


def pivot_data(data, data_type):
    """透视汇总，data_type='订单'或'售后'"""
    if data_type == '订单':
        group_keys = ['渠道', '店铺', '年份', '下单月份', 'IP', '类目', '订单区分']
        sum_fields = ['商品数量', '订单收入']
        count_field = '订单编号'
    else:
        group_keys = ['渠道', '店铺', '售后月份', '下单月份', 'IP', '类目', '订单区分']
        sum_fields = ['退商品金额']
        count_field = '订单编号'

    groups = {}
    for row in data:
        key = tuple(str(row.get(k, '')) if row.get(k) is not None else '' for k in group_keys)
        if key not in groups:
            groups[key] = {'_count': 0}
            for sf in sum_fields:
                groups[key][sf] = 0
        groups[key]['_count'] += 1
        for sf in sum_fields:
            val = row.get(sf)
            num = _to_number(val)
            if num is not None:
                groups[key][sf] += num

    int_fields = {'商品数量', '订单数量'}

    # 输出列顺序：分组键 + 求和字段（按设计规格顺序插入订单数量）
    if data_type == '订单':
        output_order = group_keys + ['商品数量', '订单数量', '订单收入']
    else:
        output_order = group_keys + ['订单数量', '退商品金额']

    result = []
    for key, agg in groups.items():
        row = dict(zip(group_keys, key))
        for sf in sum_fields:
            row[sf] = int(agg[sf]) if sf in int_fields else agg[sf]
        row['订单数量'] = agg['_count']
        # 按output_order排列
        ordered = {k: row[k] for k in output_order if k in row}
        result.append(ordered)
    return result


def pivot_merged(order_data, aftersale_data):
    """将订单和售后透视数据按(下单月份, 渠道, 店铺, IP, 合并类目)合并汇总。
    订单取下单月份，售后取售后月份；合并类目=订单区分(空则取类目)。"""
    groups = {}

    def _key(row, month_field):
        类目 = str(row.get('类目', '') if row.get('类目') is not None else '')
        订单区分 = str(row.get('订单区分', '') if row.get('订单区分') is not None else '')
        合并类目 = 订单区分 if 订单区分 else 类目
        return (str(row.get(month_field, '') if row.get(month_field) is not None else ''),
                str(row.get('渠道', '') if row.get('渠道') is not None else ''),
                str(row.get('店铺', '') if row.get('店铺') is not None else ''),
                str(row.get('IP', '') if row.get('IP') is not None else ''),
                合并类目)

    for row in order_data:
        key = _key(row, '下单月份')
        if key not in groups:
            groups[key] = {'订单收入': 0, '售后退款': 0}
        val = _to_number(row.get('订单收入'))
        if val is not None:
            groups[key]['订单收入'] += val

    for row in aftersale_data:
        key = _key(row, '售后月份')
        if key not in groups:
            groups[key] = {'订单收入': 0, '售后退款': 0}
        val = _to_number(row.get('退商品金额'))
        if val is not None:
            groups[key]['售后退款'] += val

    result = []
    for key, agg in sorted(groups.items()):
        收入 = agg['订单收入']
        退款 = agg['售后退款']
        result.append({
            '下单月份': key[0],
            '渠道': key[1],
            '店铺': key[2],
            'IP': key[3],
            '合并类目': key[4],
            'IP结算金额': round(收入 - 退款, 2),
            '订单收入': round(收入, 2),
            '售后退款': round(退款, 2),
        })
    return result


def pivot_merged_by_category(order_data, aftersale_data):
    """订单+售后按(渠道, 店铺, 月份, IP, 类目)合并汇总。
    订单取下单月份，售后取售后月份（当月收款/当月退款）。"""
    groups = {}

    def _key(row, month_field):
        return (str(row.get('渠道', '') or ''),
                str(row.get('店铺', '') or ''),
                str(row.get(month_field, '') or ''),
                str(row.get('IP', '') or ''),
                str(row.get('类目', '') or ''))

    for row in order_data:
        key = _key(row, '下单月份')
        if key not in groups:
            groups[key] = {'订单收入': 0, '售后退款': 0}
        val = _to_number(row.get('订单收入'))
        if val is not None:
            groups[key]['订单收入'] += val

    for row in aftersale_data:
        key = _key(row, '售后月份')
        if key not in groups:
            groups[key] = {'订单收入': 0, '售后退款': 0}
        val = _to_number(row.get('退商品金额'))
        if val is not None:
            groups[key]['售后退款'] += val

    output_fields = ['渠道', '店铺', '月份', 'IP', '类目']
    result = []
    for key, agg in sorted(groups.items()):
        row = dict(zip(output_fields, key))
        row['订单收入'] = round(agg['订单收入'], 2)
        row['售后退款'] = round(agg['售后退款'], 2)
        result.append(row)
    return result


def pivot_ip_summary(order_data, aftersale_data):
    """按(月份, 渠道, 店铺, IP)汇总。订单用下单月份，售后用售后月份。"""
    groups = {}
    for row in order_data:
        key = (str(row.get('下单月份', '') or ''),
               str(row.get('渠道', '') or ''),
               str(row.get('店铺', '') or ''),
               str(row.get('IP', '') or ''))
        if key not in groups:
            groups[key] = {'订单收入': 0, '退款金额': 0}
        val = _to_number(row.get('订单收入'))
        if val is not None:
            groups[key]['订单收入'] += val
    for row in aftersale_data:
        key = (str(row.get('售后月份', '') or ''),
               str(row.get('渠道', '') or ''),
               str(row.get('店铺', '') or ''),
               str(row.get('IP', '') or ''))
        if key not in groups:
            groups[key] = {'订单收入': 0, '退款金额': 0}
        val = _to_number(row.get('退商品金额'))
        if val is not None:
            groups[key]['退款金额'] += val
    result = []
    for key, agg in sorted(groups.items()):
        收入 = round(agg['订单收入'], 2)
        退款 = round(agg['退款金额'], 2)
        result.append({
            '月份': key[0], '渠道': key[1], '店铺': key[2], 'IP': key[3],
            '订单收入': 收入, '退款金额': 退款, '净收入': round(收入 - 退款, 2),
        })
    return result


# === Task 9: 未匹配商品列表 ===

def append_unmatched_to_mapping(mapping_xlsx, unmatched_list, channel_rules):
    """将未匹配商品追加到映射库对应渠道sheet末尾，供人工补充IP/类目/区分"""
    if not unmatched_list:
        return

    # 按渠道分组
    by_channel = {}
    for u in unmatched_list:
        ch = u.get('渠道', '')
        by_channel.setdefault(ch, []).append(u)

    wb = load_workbook(mapping_xlsx)

    for ch_key, items in by_channel.items():
        if ch_key not in channel_rules or ch_key.startswith('_'):
            continue
        mapping_config = channel_rules[ch_key]['映射库']
        sheet_name = mapping_config['sheet名']
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        header_row_idx = mapping_config.get('header行', 0)
        # 读header
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row_idx + 1, column=c).value
            headers.append(str(val) if val else '')

        key_cols = mapping_config['引用键']  # 映射库的列名
        match_keys = mapping_config.get('匹配键', key_cols)  # 统一字段名

        # 收集已有引用键，避免重复追加
        existing_keys = set()
        for r in range(header_row_idx + 2, ws.max_row + 1):
            parts = []
            for kc in key_cols:
                if kc in headers:
                    ci = headers.index(kc)
                    val = ws.cell(row=r, column=ci + 1).value
                    parts.append(_normalize_value(val))
            existing_keys.add(''.join(parts))

        # 追加新行
        added = 0
        for item in items:
            ref_key = item.get('引用键', '')
            if ref_key in existing_keys:
                continue
            existing_keys.add(ref_key)

            next_row = ws.max_row + 1
            # 填入引用键对应的列值
            for kc, mk in zip(key_cols, match_keys):
                if kc in headers:
                    ci = headers.index(kc)
                    ws.cell(row=next_row, column=ci + 1, value=item.get(mk, ''))
            added += 1

        if added:
            print(f'  映射库 [{sheet_name}] 追加{added}条未匹配商品')

    _save_with_retry(wb, mapping_xlsx)
    print(f'  映射库已更新: {mapping_xlsx}')


def _get_mapping_key_fields(channel_rules, channel_key, data_type):
    """获取映射匹配时用的统一字段名列表。优先用显式匹配键，否则从字段映射推导。"""
    mapping_config = channel_rules[channel_key]['映射库']
    if '匹配键' in mapping_config:
        return mapping_config['匹配键']
    field_map = channel_rules[channel_key]['字段映射'][data_type]
    ref_keys = mapping_config['引用键']
    return [field_map.get(rk, rk) for rk in ref_keys]


def process_channel(channel_key, rules, source_dir, store_mapping, mapping_xlsx, global_month_map=None):
    """处理单个渠道，返回 (订单透视list, 售后透视list, 所有未匹配list)"""
    ch_rules = rules[channel_key]
    print(f'\n=== 处理渠道: {channel_key} ===')

    # 1. 扫描文件
    files = scan_files(source_dir, channel_key, ch_rules['文件识别'])
    print(f"  订单文件: {len(files['订单'])}个, 售后文件: {len(files['售后'])}个")
    if not files['订单'] and not files['售后']:
        return [], [], [], [], []

    # 加载映射库
    mapping_config = ch_rules['映射库']
    mapping_dict = load_mapping_dict(mapping_xlsx, mapping_config)
    get_fields = mapping_config['获取字段']

    all_orders = []
    all_aftersales = []
    all_unmatched = []

    for dtype in ['订单', '售后']:
        for fpath in files[dtype]:
            fname = os.path.basename(fpath)
            print(f'  处理: {fname}')

            # 2. 读取
            data = read_file(fpath, ch_rules['读取方式'][dtype])
            print(f'    读取: {len(data)}行')

            # 3. 清除不可见字符 + 空行
            data = clean_strings(data)
            data = remove_empty_rows(data)

            # 4. 过滤
            data = apply_filter(data, ch_rules['过滤规则'][dtype])
            print(f'    过滤后: {len(data)}行')

            # 5. 字段映射
            data = map_fields(data, ch_rules['字段映射'][dtype])

            # 6. 月份提取
            data = extract_month(data, ch_rules['月份提取'][dtype])

            # 7. 映射匹配（只对订单做，售后靠补充）
            if dtype == '订单':
                key_fields = _get_mapping_key_fields(rules, channel_key, dtype)
                data, unmatched = match_products(data, mapping_dict, key_fields, get_fields)
                # 关键词兜底：映射库未匹配时用商品名关键词填充
                keyword_rules = rules.get('_商品名关键词映射', [])
                if keyword_rules:
                    still_unmatched = []
                    kw_hit = 0
                    matched_rows = {id(u): u for u in unmatched}
                    for row in data:
                        if not _not_empty(row.get('IP')):
                            name = str(row.get('商品名', '')).strip()
                            if name:
                                for kr in keyword_rules:
                                    if kr['含'] in name:
                                        for f in get_fields:
                                            if not _not_empty(row.get(f)) and f in kr:
                                                row[f] = kr[f]
                                        kw_hit += 1
                                        break
                    if kw_hit:
                        print(f'    关键词兜底匹配: {kw_hit}条')
                    # 重新判定未匹配（关键词命中的不再算未匹配）
                    for u in unmatched:
                        name = str(u.get('商品名', u.get(key_fields[0], ''))).strip() if key_fields else ''
                        hit = False
                        if name:
                            for kr in keyword_rules:
                                if kr['含'] in name:
                                    hit = True
                                    break
                        if not hit:
                            still_unmatched.append(u)
                    unmatched = still_unmatched
                for u in unmatched:
                    u['渠道'] = channel_key
                    u['文件'] = fname
                all_unmatched.extend(unmatched)

            # 添加渠道/店铺信息
            channel_display, store_name = extract_store_info(fname, channel_key, store_mapping)
            for row in data:
                row['渠道'] = channel_display
                row['店铺'] = store_name
                if dtype == '订单':
                    row['年份'] = str(row.get('下单月份', ''))[:4]

            if dtype == '订单':
                all_orders.extend(data)
            else:
                all_aftersales.extend(data)

    # 8. 售后补充（订单编号匹配）
    if all_aftersales:
        all_aftersales = supplement_aftersale(all_aftersales, all_orders, global_month_map)

    # 9. 售后商品名兜底匹配（订单编号匹配不到IP时，用商品名查映射库）
    if all_aftersales:
        name_map = build_name_mapping(mapping_xlsx, mapping_config)
        keyword_rules = rules.get('_商品名关键词映射', [])
        fallback_count = 0
        keyword_count = 0
        for row in all_aftersales:
            if not _not_empty(row.get('IP')):
                name = str(row.get('商品名', '')).strip()
                # 商品名精确匹配
                if name and name in name_map:
                    matched = name_map[name]
                    for f in get_fields:
                        if not _not_empty(row.get(f)):
                            row[f] = matched.get(f, '')
                    if _not_empty(row.get('IP')):
                        fallback_count += 1
                        continue
                # 关键词匹配
                if name:
                    for kr in keyword_rules:
                        if kr['含'] in name:
                            for f in get_fields:
                                if not _not_empty(row.get(f)) and f in kr:
                                    row[f] = kr[f]
                            keyword_count += 1
                            break
        if fallback_count:
            print(f'  售后商品名兜底匹配: {fallback_count}条')
        if keyword_count:
            print(f'  售后关键词兜底匹配: {keyword_count}条')

    # 10. 透视
    order_pivot = pivot_data(all_orders, '订单') if all_orders else []
    aftersale_pivot = pivot_data(all_aftersales, '售后') if all_aftersales else []

    print(f'  订单透视: {len(order_pivot)}行, 售后透视: {len(aftersale_pivot)}行')
    if all_unmatched:
        print(f'  未匹配商品: {len(all_unmatched)}条')
    return order_pivot, aftersale_pivot, all_unmatched, all_orders, all_aftersales


def _extract_detail(data, fields):
    """从明细数据中提取指定字段列表"""
    result = []
    for row in data:
        r = {}
        for f in fields:
            r[f] = row.get(f, '')
        result.append(r)
    return result


def _dedup_detail(data, id_fields, priority_fields):
    """按唯一标识去重，同标识多条时保留关键字段最完整的（同等完整度取后出现的）。
    id_fields: 构成唯一键的字段列表；priority_fields: 用于比较完整度的字段列表。"""
    if not data:
        return data, 0
    best_idx = {}  # key → (index, score)
    for i, row in enumerate(data):
        key = tuple(str(row.get(f, '')).strip() for f in id_fields)
        if not any(key):
            continue
        score = sum(1 for f in priority_fields if _not_empty(row.get(f)))
        if key not in best_idx:
            best_idx[key] = (i, score)
        else:
            _, old_score = best_idx[key]
            if score >= old_score:
                best_idx[key] = (i, score)
    keep = set(idx for idx, _ in best_idx.values())
    seen_keys = set()
    result = []
    removed = 0
    for i, row in enumerate(data):
        key = tuple(str(row.get(f, '')).strip() for f in id_fields)
        if not any(key):
            result.append(row)
        elif i in keep:
            result.append(row)
        else:
            removed += 1
    return result, removed


def _resolve_source_dirs(raw_paths):
    """解析数据源路径：含YYYYMM店铺子文件夹时自动展开，否则直接使用"""
    source_dirs = []
    for p in raw_paths:
        p = p.strip().strip('"').strip("'").rstrip('\\')
        if not p or not os.path.isdir(p):
            if p:
                print(f'  路径跳过(不存在): {p}')
            continue
        subs = sorted([
            os.path.join(p, d) for d in os.listdir(p)
            if os.path.isdir(os.path.join(p, d)) and re.match(r'^\d{6}店铺$', d)
        ])
        if subs:
            source_dirs.extend(subs)
            print(f'  自动展开: {os.path.basename(p)} → {", ".join(os.path.basename(s) for s in subs)}')
        else:
            source_dirs.append(p)
    return source_dirs


def main(input_paths=None):
    """主入口。input_paths为None时交互输入路径，否则直接使用传入路径。全渠道处理。"""
    rules = load_channel_rules()
    store_mapping = load_store_mapping()
    mapping_xlsx = os.path.join(SCRIPT_DIR, '商品映射库.xlsx')
    channels = [k for k in rules if not k.startswith('_')]
    selected = channels

    if input_paths:
        print(f'\n全渠道处理({len(channels)}个渠道)')
        for p in input_paths:
            print(f'  输入路径: {p}')
        source_dirs = _resolve_source_dirs(input_paths)
        save_cache(input_paths)
    elif len(sys.argv) > 1:
        print(f'\n拖放模式: 全渠道处理({len(channels)}个渠道)')
        for p in sys.argv[1:]:
            print(f'  输入路径: {p}')
        source_dirs = _resolve_source_dirs(sys.argv[1:])
        save_cache(sys.argv[1:])
    else:
        paths = get_paths_interactive()
        if not paths:
            return
        save_cache(paths)
        source_dirs = _resolve_source_dirs(paths)

    if not source_dirs:
        print('无有效路径')
        return

    output_dir = OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    # 加载全域数据的订单月份映射（售后下单月份兜底查找）
    global_data_path = os.path.join(SCRIPT_DIR, '公域数据最新.xlsx')
    global_month_map = load_global_order_months(global_data_path)

    all_unmatched = []
    all_order_detail = []
    all_aftersale_detail = []

    for source_dir in source_dirs:
        print(f'\n>>> 数据源: {os.path.basename(source_dir)}')
        for ch in selected:
            _, _, unmatched, raw_orders, raw_aftersales = process_channel(
                ch, rules, source_dir, store_mapping, mapping_xlsx, global_month_map)
            all_unmatched.extend(unmatched)
            all_order_detail.extend(raw_orders)
            all_aftersale_detail.extend(raw_aftersales)

    # 跨目录售后二次补充：用全部订单数据再次填充仍缺失的售后字段
    if all_aftersale_detail:
        full_order_map = {}
        for o in all_order_detail:
            oid = _normalize_oid(o.get('订单编号', ''))
            if oid:
                full_order_map.setdefault(oid, []).append(o)
        cross_fix = 0
        for row in all_aftersale_detail:
            need_fix = not _not_empty(row.get('下单月份')) or not _not_empty(row.get('IP'))
            if not need_fix:
                continue
            oid = _normalize_oid(row.get('订单编号', ''))
            candidates = full_order_map.get(oid, [])
            if candidates:
                matched = _best_order_match(row, candidates)
                for f in ['IP', '类目', '订单区分', '下单月份']:
                    if not _not_empty(row.get(f)):
                        row[f] = matched.get(f, '')
                cross_fix += 1
            elif not _not_empty(row.get('下单月份')) and global_month_map and oid:
                row['下单月份'] = global_month_map.get(oid, '')
                if _not_empty(row.get('下单月份')):
                    cross_fix += 1
        if cross_fix:
            print(f'\n  跨目录售后补充: {cross_fix}条')

    # 提取当前批次月份（从目录名 YYYYMM店铺）
    batch_months = set()
    for sd in source_dirs:
        m = re.match(r'^(\d{6})店铺$', os.path.basename(sd))
        if m:
            batch_months.add(m.group(1))
    if batch_months:
        print(f'\n当前批次月份: {", ".join(sorted(batch_months))}')

    # 提取当前批次明细
    order_fields = ['渠道', '店铺', '年份', '下单月份', 'IP', '类目', '订单区分',
                    '商品名', '商品数量', '订单收入', '订单编号']
    aftersale_fields = ['渠道', '店铺', '售后月份', '下单月份', 'IP', '类目', '订单区分',
                        '商品名', '退商品金额', '订单编号', '售后编号']
    new_order_detail = _extract_detail(all_order_detail, order_fields)
    new_aftersale_detail = _extract_detail(all_aftersale_detail, aftersale_fields)

    # 存档合并：加载已有源数据，按批次月份去重后合并
    detail_path = os.path.join(output_dir, '源数据_公域订单.xlsx')
    existing = read_output_sheets(detail_path)
    old_orders = existing.get('订单明细', [])
    old_aftersales = existing.get('售后明细', [])

    # 清理存档脏数据（误扫描的非订单文件产生的行，如"账单"类文件）
    def _is_dirty(row):
        store = str(row.get('店铺', ''))
        return '账单' in store
    old_orders = [r for r in old_orders if not _is_dirty(r)]
    old_aftersales = [r for r in old_aftersales if not _is_dirty(r)]

    if batch_months and (old_orders or old_aftersales):
        old_count_o, old_count_a = len(old_orders), len(old_aftersales)
        old_orders = [r for r in old_orders if _parse_yyyymm(r.get('下单月份')) not in batch_months]
        old_aftersales = [r for r in old_aftersales if _parse_yyyymm(r.get('售后月份')) not in batch_months]
        print(f'  存档: 旧订单{old_count_o}→{len(old_orders)}行(去除批次月份), 新订单+{len(new_order_detail)}行')
        print(f'  存档: 旧售后{old_count_a}→{len(old_aftersales)}行(去除批次月份), 新售后+{len(new_aftersale_detail)}行')

    merged_orders = old_orders + new_order_detail
    merged_aftersales = old_aftersales + new_aftersale_detail

    # 去重：跨批次合并可能产生重复记录（售后/订单文件含跨月数据时）
    merged_orders, dedup_o = _dedup_detail(
        merged_orders, ['订单编号', '商品名'], ['下单月份', 'IP', '类目'])
    merged_aftersales, dedup_a = _dedup_detail(
        merged_aftersales, ['售后编号'], ['下单月份', 'IP', '类目'])
    if dedup_o or dedup_a:
        print(f'  去重: 订单-{dedup_o}条, 售后-{dedup_a}条')

    # 合并后二次补充：用merged_orders(含历史最佳数据)修复仍缺IP的售后
    if merged_aftersales:
        merge_order_map = {}
        for o in merged_orders:
            oid = _normalize_oid(o.get('订单编号', ''))
            if oid:
                merge_order_map.setdefault(oid, []).append(o)
        keyword_rules = rules.get('_商品名关键词映射', [])
        merge_fix = 0
        for row in merged_aftersales:
            if _not_empty(row.get('IP')):
                continue
            oid = _normalize_oid(row.get('订单编号', ''))
            # 订单编号匹配（优先有IP的候选）
            candidates = merge_order_map.get(oid, [])
            if candidates:
                matched = _best_order_match(row, candidates)
                if _not_empty(matched.get('IP')):
                    for f in ['IP', '类目', '订单区分', '下单月份']:
                        if not _not_empty(row.get(f)):
                            row[f] = matched.get(f, '')
                    merge_fix += 1
                    continue
            # 关键词兜底
            name = str(row.get('商品名', '')).strip()
            if name and keyword_rules:
                for kr in keyword_rules:
                    if kr['含'] in name:
                        for f in ['IP', '类目', '订单区分']:
                            if not _not_empty(row.get(f)) and f in kr:
                                row[f] = kr[f]
                        if _not_empty(row.get('IP')):
                            merge_fix += 1
                        break
        if merge_fix:
            print(f'  合并后二次补充: {merge_fix}条')

    # 从全量明细重建所有pivot sheet
    sheets = {}
    order_by_channel = {}
    for row in merged_orders:
        ch = str(row.get('渠道') or '')
        order_by_channel.setdefault(ch, []).append(row)
    aftersale_by_channel = {}
    for row in merged_aftersales:
        ch = str(row.get('渠道') or '')
        aftersale_by_channel.setdefault(ch, []).append(row)

    all_order_pivot = []
    all_aftersale_pivot = []
    for ch in sorted(set(list(order_by_channel.keys()) + list(aftersale_by_channel.keys()))):
        if order_by_channel.get(ch):
            op = pivot_data(order_by_channel[ch], '订单')
            sheets[f'{ch}_订单'] = op
            all_order_pivot.extend(op)
        if aftersale_by_channel.get(ch):
            ap = pivot_data(aftersale_by_channel[ch], '售后')
            sheets[f'{ch}_售后'] = ap
            all_aftersale_pivot.extend(ap)

    if all_order_pivot:
        sheets['汇总_订单'] = all_order_pivot
    if all_aftersale_pivot:
        sheets['汇总_售后'] = all_aftersale_pivot

    if merged_orders or merged_aftersales:
        sheets['汇总合并'] = pivot_merged(merged_orders, merged_aftersales)
        sheets['汇总合并（类目）'] = pivot_merged_by_category(merged_orders, merged_aftersales)
        sheets['IP汇总'] = pivot_ip_summary(merged_orders, merged_aftersales)

    # 未匹配商品列表（去重）
    unique_unmatched = []
    if all_unmatched:
        seen = set()
        for u in all_unmatched:
            key = u['引用键']
            if key not in seen:
                seen.add(key)
                unique_unmatched.append(u)
        sheets['未匹配商品'] = unique_unmatched
        print(f'\n未匹配商品: {len(unique_unmatched)}种（去重后）')

    output_path = os.path.join(output_dir, '清洗_公域订单.xlsx')
    write_output(sheets, output_path)

    # 源数据（全量合并明细，作为持久存档）
    detail_sheets = {}
    if merged_orders:
        detail_sheets['订单明细'] = merged_orders
    if merged_aftersales:
        detail_sheets['售后明细'] = merged_aftersales
    if detail_sheets:
        write_output(detail_sheets, detail_path)
        print(f'  源数据(存档): {detail_path}')

    # 未匹配商品写回映射库
    if all_unmatched:
        append_unmatched_to_mapping(mapping_xlsx, unique_unmatched, rules)

    # 更新本地订单月份映射缓存（每月新订单自动累积）
    update_order_month_cache(all_order_detail)

    # 汇总日志
    order_income = sum(_to_number(r.get('订单收入')) or 0 for r in merged_orders)
    refund_amount = sum(_to_number(r.get('退商品金额')) or 0 for r in merged_aftersales)
    print('\n' + '=' * 50)
    print('处理完成!')
    print(f'  渠道数: {len(selected)}')
    print(f'  数据源: {len(source_dirs)}个目录')
    print(f'  本次处理: 订单{len(new_order_detail)}行, 售后{len(new_aftersale_detail)}行')
    print(f'  全量合计: 订单{len(merged_orders)}行, 售后{len(merged_aftersales)}行')
    print(f'  订单收入: {order_income:.2f}, 退款: {refund_amount:.2f}')
    if unique_unmatched:
        print(f'  未匹配商品: {len(unique_unmatched)}种')
    print(f'  输出: {output_path}')
    print('=' * 50)


if __name__ == '__main__':
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    print('=' * 50)
    print('  数据清洗工具')
    print(f'  工作目录: {SCRIPT_DIR}')
    print('=' * 50)
    try:
        main()
    except Exception as e:
        print(f'\n*** 运行出错: {e} ***')
        import traceback
        traceback.print_exc()
    input('\n按回车键退出...')
