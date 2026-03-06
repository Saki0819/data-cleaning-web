"""支付宝/余利宝账单清洗模块"""
import os
import sys
import json
import zipfile
import tempfile
import shutil
import calendar
from collections import defaultdict
from openpyxl import load_workbook

# === 路径机制（与其他模块一致） ===
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.join(os.path.dirname(sys.executable), 'clean_alipay')
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'output'))

sys.path.insert(0, os.path.normpath(os.path.join(SCRIPT_DIR, '..')))
from utils import clean_strings, remove_empty_rows, _parse_date, _to_number, write_output, read_xml_spreadsheet, read_output_sheets

# === 配置 ===
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'alipay_config.json')


def _load_config():
    """加载alipay_config.json"""
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def _load_product_map(config):
    """加载商品映射表，返回 {商品名称: {'IP': ..., '小阶初阶': ...}}"""
    xlsx_path = os.path.join(SCRIPT_DIR, config['商品映射路径'])
    if not os.path.exists(xlsx_path):
        print(f'*** 商品映射表不存在: {xlsx_path} ***')
        return {}
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ''
        if not name:
            continue
        ip = str(row[1]).strip() if row[1] else ''
        belong = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        result[name] = {'IP': ip, '小阶初阶': belong}
    wb.close()
    return result


def _scan_files(input_path):
    """扫描目录及直属子目录，返回(zip列表, 余利宝xlsx列表)"""
    input_path = input_path.strip().strip('"').strip("'").rstrip('\\').rstrip('/')
    if not os.path.isdir(input_path):
        print(f'路径不存在: {input_path}')
        return [], []

    # 搜索当前目录 + 直属子目录
    search_dirs = [input_path]
    for name in os.listdir(input_path):
        sub = os.path.join(input_path, name)
        if os.path.isdir(sub):
            search_dirs.append(sub)

    zips = []
    ylb_files = []
    for d in search_dirs:
        for f in sorted(os.listdir(d)):
            fpath = os.path.join(d, f)
            if not os.path.isfile(fpath):
                continue
            if f.lower().endswith('.zip'):
                zips.append(fpath)
            elif f.lower().endswith('.xlsx') and '余利宝' in f and not f.startswith('~$'):
                ylb_files.append(fpath)

    return zips, ylb_files


def _extract_merchant_id(zip_filename, merchant_map):
    """从zip文件名提取商户ID，返回(商户ID, 店铺名)或(None, None)"""
    basename = os.path.basename(zip_filename)
    for mid, shop in merchant_map.items():
        if mid in basename:
            return mid, shop
    return None, None


def _month_last_day(yyyymm):
    """给定yyyymm字符串，返回该月最后一天 yyyy-mm-dd"""
    if not yyyymm or len(yyyymm) < 6:
        return ''
    try:
        year = int(yyyymm[:4])
        month = int(yyyymm[4:6])
        _, last_day = calendar.monthrange(year, month)
        return f'{year:04d}-{month:02d}-{last_day:02d}'
    except (ValueError, TypeError):
        return ''


# ========== 支付宝账单处理 ==========

def process_zip(zip_path, config, product_map):
    """处理单个支付宝zip文件，返回(data_rows, unmatched_products)"""
    merchant_map = config['商户映射']
    type_map = config['账务类型映射']

    mid, shop = _extract_merchant_id(zip_path, merchant_map)
    if not mid:
        print(f'  *** 文件名未匹配任何商户ID，跳过: {os.path.basename(zip_path)} ***')
        return [], set()

    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(tmpdir)
            names = zf.namelist()

        all_rows = []
        unmatched = set()

        for name in names:
            fpath = os.path.join(tmpdir, name)
            if not os.path.isfile(fpath):
                continue

            data = read_xml_spreadsheet(fpath)
            clean_strings(data)
            remove_empty_rows(data)
            print(f'  {os.path.basename(zip_path)} → {len(data)}行')

            for row in data:
                # 跳过无入账时间的无效行（XML尾部空行）
                if not row.get('入账时间', '').strip():
                    continue

                # 收入/支出转float
                row['收入（+元）'] = _to_number(row.get('收入（+元）', '')) or 0.0
                row['支出（-元）'] = _to_number(row.get('支出（-元）', '')) or 0.0

                # 店铺
                row['店铺'] = shop

                # 账务类型→明细/归类
                biz_type = row.get('账务类型', '').strip()
                mapping = type_map.get(biz_type, {'明细': biz_type, '归类': '其他'})
                row['明细'] = mapping['明细']
                row['归类'] = mapping['归类']

                # 时间提取
                月份, 日期 = _parse_date(row.get('入账时间', ''))
                row['月份'] = 月份 or ''
                row['日期'] = 日期 if 日期 else ''

                # 商品映射 + IP规则
                product_name = row.get('商品名称', '').strip()
                pm = product_map.get(product_name)
                if row['明细'] == '提现':
                    # 提现：IP固定"其他"
                    row['IP'] = '其他'
                    row['小阶初阶'] = ''
                elif pm:
                    row['IP'] = pm['IP']
                    row['小阶初阶'] = pm['小阶初阶']
                elif row['归类'] == '余利宝':
                    # 余利宝未匹配：IP默认"其他"
                    row['IP'] = '其他'
                    row['小阶初阶'] = ''
                else:
                    row['IP'] = ''
                    row['小阶初阶'] = ''
                    if product_name:
                        unmatched.add(product_name)

                all_rows.append(row)

        return all_rows, unmatched
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ========== 余利宝账单处理 ==========

def process_yulebao(xlsx_path, config):
    """处理单个余利宝xlsx文件，返回data_rows列表

    余利宝xlsx结构：
    Row1: 标题 "余利宝明细对账单"
    Row2: "余利宝账户名称" + 公司名（含关键词匹配店铺）
    Row3: "余利宝账号" + 账号
    Row4: 列名（交易时间/交易类型/交易金额/余额/交易名称/对方户名/对方账号/备注）
    Row5+: 数据
    """
    account_map = config.get('余利宝账户映射', {})
    type_map = config.get('余利宝交易类型映射', {})

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 从Row2提取账户名称→匹配店铺
    account_name = str(ws.cell(2, 2).value or '').strip()
    shop = ''
    for keyword, shop_name in account_map.items():
        if keyword in account_name:
            shop = shop_name
            break
    if not shop:
        shop = account_name
        print(f'  *** 未匹配余利宝账户: {account_name}，使用原名 ***')

    # Row4为列名，Row5+为数据
    headers = []
    for j in range(1, ws.max_column + 1):
        h = ws.cell(4, j).value
        headers.append(str(h).strip().replace('\n', '') if h else '')

    all_rows = []
    for i in range(5, ws.max_row + 1):
        row = {}
        for j, h in enumerate(headers):
            val = ws.cell(i, j + 1).value
            row[h] = str(val).strip() if val is not None else ''
        if not row.get('交易时间', '').strip():
            continue

        # 交易类型→明细
        tx_type = row.get('交易类型', '').strip()
        row['明细'] = type_map.get(tx_type, tx_type)

        # 时间提取
        月份, 日期 = _parse_date(row['交易时间'])
        row['月份'] = 月份 or ''
        row['日期'] = 日期 if 日期 else ''

        # 固定字段
        row['店铺'] = shop
        row['IP'] = '其他'
        row['渠道'] = '支付宝'

        # 交易金额→收入/支出
        amount_str = row.get('交易金额', '').strip()
        if amount_str.startswith('+'):
            row['收入'] = _to_number(amount_str[1:]) or 0.0
            row['支出'] = 0.0
        elif amount_str.startswith('-'):
            row['收入'] = 0.0
            row['支出'] = _to_number(amount_str[1:]) or 0.0
        else:
            row['收入'] = _to_number(amount_str) or 0.0
            row['支出'] = 0.0

        all_rows.append(row)

    wb.close()
    print(f'  {os.path.basename(xlsx_path)} ({account_name}) → {len(all_rows)}行')
    return all_rows


# ========== 汇总函数 ==========

def aggregate_orders(rows):
    """订单汇总（仅支付宝）：按 月份+归类+IP+店铺+小阶初阶 分组"""
    groups = defaultdict(lambda: {'收入': 0.0, '支出': 0.0})
    keys = ['月份', '归类', 'IP', '店铺', '小阶初阶']
    for row in rows:
        key = tuple(row.get(k, '') for k in keys)
        groups[key]['收入'] += row.get('收入（+元）', 0.0)
        groups[key]['支出'] += row.get('支出（-元）', 0.0)

    result = []
    for key, agg in groups.items():
        r = dict(zip(keys, key))
        r['收入'] = round(agg['收入'], 2)
        r['支出'] = round(agg['支出'], 2)
        result.append(r)
    result.sort(key=lambda x: (x['月份'], x['归类'], x['IP'], x['店铺']))
    return result


def aggregate_funds(alipay_rows, yulebao_rows):
    """资金汇总（支付宝+余利宝合并）：提现按日、其他按月（日期=月末）"""
    groups = defaultdict(lambda: {'收入': 0.0, '支出': 0.0})
    key_uses_date = {}  # key → True表示用的是实际日期，False表示用的是月份

    def _make_key(row, is_withdraw):
        if is_withdraw:
            return (row.get('日期', ''), row.get('明细', ''), row.get('IP', ''), row.get('店铺', ''))
        else:
            return (row.get('月份', ''), row.get('明细', ''), row.get('IP', ''), row.get('店铺', ''))

    # 支付宝数据
    for row in alipay_rows:
        is_w = row.get('明细', '') == '提现'
        key = _make_key(row, is_w)
        groups[key]['收入'] += row.get('收入（+元）', 0.0)
        groups[key]['支出'] += row.get('支出（-元）', 0.0)
        key_uses_date[key] = is_w

    # 余利宝数据
    for row in yulebao_rows:
        is_w = row.get('明细', '') == '提现'
        key = _make_key(row, is_w)
        groups[key]['收入'] += row.get('收入', 0.0)
        groups[key]['支出'] += row.get('支出', 0.0)
        key_uses_date[key] = is_w

    result = []
    for key, agg in groups.items():
        date_or_month, 明细, ip, shop = key
        r = {'明细': 明细, 'IP': ip, '店铺': shop}
        if key_uses_date.get(key):
            r['日期'] = date_or_month  # 提现：实际日期
        else:
            r['日期'] = _month_last_day(date_or_month)  # 非提现：月末
        r['收入'] = round(agg['收入'], 2)
        r['支出'] = round(agg['支出'], 2)
        result.append(r)
    result.sort(key=lambda x: (x.get('日期', ''), x['明细'], x['IP'], x['店铺']))
    return result


# 输出列顺序
ALIPAY_EXTRA_COLUMNS = ['店铺', '明细', '归类', '月份', '日期', 'IP', '小阶初阶']
YLB_EXTRA_COLUMNS = ['店铺', '明细', '月份', '日期', 'IP', '渠道', '收入', '支出']
ORDER_SUMMARY_COLUMNS = ['月份', '归类', 'IP', '店铺', '小阶初阶', '收入', '支出']
FUND_SUMMARY_COLUMNS = ['日期', '明细', 'IP', '店铺', '收入', '支出']


def _reorder(rows, columns):
    """按指定列顺序重新排列dict"""
    return [{col: r.get(col, '') for col in columns} for r in rows]


def main(input_paths):
    """主入口，接收路径列表"""
    config = _load_config()
    product_map = _load_product_map(config)
    print(f'商品映射: {len(product_map)}条')

    input_path = input_paths[0] if input_paths else ''
    if not input_path:
        print('未提供路径')
        return

    # 扫描文件（支持目录及子目录）
    print(f'\n扫描目录: {input_path}')
    zip_files, ylb_files = _scan_files(input_path)

    if not zip_files and not ylb_files:
        print('未找到支付宝zip或余利宝xlsx文件')
        return

    # === 处理支付宝账单 ===
    alipay_data = []
    all_unmatched = set()
    if zip_files:
        print(f'\n--- 支付宝账单 ({len(zip_files)}个zip) ---')
        for zf in zip_files:
            data, unmatched = process_zip(zf, config, product_map)
            alipay_data.extend(data)
            all_unmatched.update(unmatched)
        print(f'支付宝: {len(alipay_data)}条')

    # === 处理余利宝账单 ===
    yulebao_data = []
    if ylb_files:
        print(f'\n--- 余利宝账单 ({len(ylb_files)}个文件) ---')
        for yf in ylb_files:
            data = process_yulebao(yf, config)
            yulebao_data.extend(data)
        print(f'余利宝: {len(yulebao_data)}条')

    if not alipay_data and not yulebao_data:
        print('无有效数据')
        return

    # === 提取批次月份 ===
    alipay_batch_months = set(r['月份'] for r in alipay_data if r.get('月份'))
    ylb_batch_months = set(r['月份'] for r in yulebao_data if r.get('月份'))
    all_batch_months = alipay_batch_months | ylb_batch_months
    if all_batch_months:
        print(f'当前批次月份: {", ".join(sorted(all_batch_months))}')

    # === 存档合并 ===
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, '支付宝账单清洗.xlsx')
    existing = read_output_sheets(output_path)
    old_alipay = existing.get('支付宝_源数据', [])
    old_yulebao = existing.get('余利宝_源数据', [])

    # 旧数据修正（xlsx读回None/int→''/float，日期格式统一）
    for r in old_alipay:
        for k in r:
            if r[k] is None:
                r[k] = ''
        r['收入（+元）'] = float(r['收入（+元）'] or 0)
        r['支出（-元）'] = float(r['支出（-元）'] or 0)
        if isinstance(r.get('日期'), str) and '/' in r['日期']:
            r['日期'] = r['日期'].replace('/', '-')
    for r in old_yulebao:
        for k in r:
            if r[k] is None:
                r[k] = ''
        r['收入'] = float(r['收入'] or 0)
        r['支出'] = float(r['支出'] or 0)
        if isinstance(r.get('日期'), str) and '/' in r['日期']:
            r['日期'] = r['日期'].replace('/', '-')

    # 按批次月份过滤旧数据
    if alipay_batch_months and old_alipay:
        old_count = len(old_alipay)
        old_alipay = [r for r in old_alipay if r.get('月份', '') not in alipay_batch_months]
        print(f'  存档: 旧支付宝{old_count}→{len(old_alipay)}行')
    if ylb_batch_months and old_yulebao:
        old_count = len(old_yulebao)
        old_yulebao = [r for r in old_yulebao if r.get('月份', '') not in ylb_batch_months]
        print(f'  存档: 旧余利宝{old_count}→{len(old_yulebao)}行')

    # 合并源数据
    merged_alipay = old_alipay + alipay_data
    merged_yulebao = old_yulebao + yulebao_data

    # === 构建输出 ===
    sheets = {}

    # 支付宝源数据
    if merged_alipay:
        ref = alipay_data[0] if alipay_data else merged_alipay[0]
        orig_cols = [k for k in ref.keys() if k not in ALIPAY_EXTRA_COLUMNS]
        sheets['支付宝_源数据'] = _reorder(merged_alipay, orig_cols + ALIPAY_EXTRA_COLUMNS)

    # 订单汇总（从全量源数据重建）
    if merged_alipay:
        order_summary = aggregate_orders(merged_alipay)
        sheets['订单汇总'] = _reorder(order_summary, ORDER_SUMMARY_COLUMNS)
        print(f'订单汇总: {len(order_summary)}行')

    # 资金汇总（从全量源数据重建）
    fund_summary = aggregate_funds(merged_alipay, merged_yulebao)
    sheets['资金汇总'] = _reorder(fund_summary, FUND_SUMMARY_COLUMNS)
    print(f'资金汇总: {len(fund_summary)}行')

    # 余利宝源数据
    if merged_yulebao:
        ref = yulebao_data[0] if yulebao_data else merged_yulebao[0]
        ylb_orig_cols = [k for k in ref.keys() if k not in YLB_EXTRA_COLUMNS]
        sheets['余利宝_源数据'] = _reorder(merged_yulebao, ylb_orig_cols + YLB_EXTRA_COLUMNS)

    # 未匹配商品（仅当前批次，不累积）
    if all_unmatched:
        sheets['未匹配商品'] = [{'商品名称': name} for name in sorted(all_unmatched)]
        print(f'未匹配商品: {len(all_unmatched)}个')

    write_output(sheets, output_path)

    # 统计（全量）
    alipay_income = sum(float(r.get('收入（+元）') or 0) for r in merged_alipay)
    alipay_expense = sum(float(r.get('支出（-元）') or 0) for r in merged_alipay)
    ylb_income = sum(float(r.get('收入') or 0) for r in merged_yulebao)
    ylb_expense = sum(float(r.get('支出') or 0) for r in merged_yulebao)
    print(f'\n处理完成!')
    if merged_alipay:
        print(f'  支付宝(全量): {len(merged_alipay)}条, 收入={alipay_income:.2f}, 支出={alipay_expense:.2f}')
    if merged_yulebao:
        print(f'  余利宝(全量): {len(merged_yulebao)}条, 收入={ylb_income:.2f}, 支出={ylb_expense:.2f}')
    print(f'  输出: {output_path}')


if __name__ == '__main__':
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    try:
        main(sys.argv[1:] or None)
    except Exception as e:
        print(f'*** 出错: {e} ***')
        import traceback
        traceback.print_exc()
    input('\n按回车退出...')
