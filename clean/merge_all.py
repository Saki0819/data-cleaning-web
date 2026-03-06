"""总合并 — 将各模块输出汇总合并为一个xlsx"""
import os, sys

if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
sys.path.insert(0, SCRIPT_DIR)
from utils import write_output


def _read_sheet(filepath, sheet_name):
    """读取xlsx指定sheet → list of dict"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h) if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        d = {}
        for i, h in enumerate(headers):
            if h:
                val = row[i] if i < len(row) else ''
                d[h] = val if val is not None else ''
        result.append(d)
    return result


def _copy_sheets(filepath, filename, mapping, sheets):
    """从文件复制指定sheet，mapping: {源sheet名: 目标sheet名}"""
    if not os.path.exists(filepath):
        print(f'  跳过（文件不存在）: {filename}')
        return
    print(f'读取: {filename}')
    for src_name, dst_name in mapping.items():
        data = _read_sheet(filepath, src_name)
        if data:
            sheets[dst_name] = data
            print(f'  {dst_name}: {len(data)}行')


def _build_fund_summary(wechat_file, alipay_file, bill_file, qiwei_file, bank_file):
    """合并微信+支付宝+公域账单+企微+银行的资金汇总"""
    fund_rows = []

    def _fund_row(date='', detail='', ip='', channel='', store='', income=0, expense=0, summary='', payee=''):
        return {'日期': date or '', '明细': detail or '', 'IP': ip or '', '渠道': channel or '', '店铺': store or '',
                '摘要': summary or '', '收(付)方名称': payee or '', '收入': income or 0, '支出': expense or 0}

    # 微信：提现汇总 + 不含提现汇总
    if os.path.exists(wechat_file):
        for sheet_name in ['提现汇总', '不含提现汇总']:
            data = _read_sheet(wechat_file, sheet_name)
            for r in data:
                fund_rows.append(_fund_row(
                    r.get('日期', ''), r.get('明细', ''), r.get('IP', ''),
                    r.get('渠道', ''), r.get('店铺', r.get('账号名称', '')),
                    r.get('日收入', 0), r.get('日支出', 0)))
            if data:
                print(f'  微信_{sheet_name}: {len(data)}行')

    # 支付宝资金汇总
    if os.path.exists(alipay_file):
        data = _read_sheet(alipay_file, '资金汇总')
        for r in data:
            fund_rows.append(_fund_row(
                r.get('日期', ''), r.get('明细', ''), r.get('IP', ''),
                '支付宝', r.get('店铺', ''), r.get('收入', 0), r.get('支出', 0)))
        if data:
            print(f'  支付宝_资金汇总: {len(data)}行')

    # 公域账单资金汇总
    if os.path.exists(bill_file):
        data = _read_sheet(bill_file, '资金汇总')
        for r in data:
            fund_rows.append(_fund_row(
                r.get('结算日期', ''), r.get('明细', ''), r.get('IP', ''),
                r.get('渠道', ''), r.get('店铺', ''), r.get('收入', 0), r.get('支出', 0)))
        if data:
            print(f'  公域账单_资金汇总: {len(data)}行')

    # 企微资金汇总
    if os.path.exists(qiwei_file):
        data = _read_sheet(qiwei_file, '资金汇总')
        for r in data:
            fund_rows.append(_fund_row(
                r.get('结算日期', ''), r.get('明细', ''), r.get('IP', ''),
                r.get('渠道', ''), r.get('店铺', ''), r.get('收入', 0), r.get('支出', 0)))
        if data:
            print(f'  企微_资金汇总: {len(data)}行')

    # 银行资金汇总（含摘要和收付方）
    if os.path.exists(bank_file):
        data = _read_sheet(bank_file, '资金汇总')
        for r in data:
            fund_rows.append(_fund_row(
                r.get('日期', ''), r.get('明细', ''), r.get('IP', ''),
                r.get('渠道', ''), r.get('店铺', ''), r.get('收入', 0), r.get('支出', 0),
                r.get('摘要', ''), r.get('收(付)方名称', '')))
        if data:
            print(f'  银行_资金汇总: {len(data)}行')

    return fund_rows


def main():
    print('\n>>> 启动总合并...')

    shop_file = os.path.join(OUTPUT_DIR, '清洗_公域订单.xlsx')
    crm_file = os.path.join(OUTPUT_DIR, '清洗_私域订单.xlsx')
    wechat_file = os.path.join(OUTPUT_DIR, '微信账单清洗.xlsx')
    alipay_file = os.path.join(OUTPUT_DIR, '支付宝账单清洗.xlsx')
    bill_file = os.path.join(OUTPUT_DIR, '公域账单清洗.xlsx')
    qiwei_file = os.path.join(OUTPUT_DIR, '企微账单清洗.xlsx')
    bank_file = os.path.join(OUTPUT_DIR, '银行账单清洗.xlsx')

    sheets = {}

    # 1. shop → 公域_ 前缀
    _copy_sheets(shop_file, '清洗_公域订单.xlsx', {
        '汇总合并': '公域_汇总合并',
        '汇总合并（类目）': '公域_汇总合并（类目）',
        'IP汇总': '公域_IP汇总',
    }, sheets)

    # 2. crm → CRM_ 前缀
    _copy_sheets(crm_file, '清洗_私域订单.xlsx', {
        '订单汇总': 'CRM_订单汇总',
        '退款汇总': 'CRM_退款汇总',
        '合并汇总': 'CRM_合并汇总',
        'IP汇总': 'CRM_IP汇总',
    }, sheets)

    # 3. alipay 订单汇��� → 支付宝_ 前缀
    _copy_sheets(alipay_file, '支付宝账单清洗.xlsx', {
        '订单汇总': '支付宝_订单汇总',
    }, sheets)

    # 4. 资金汇总（微信+支付宝+公域账单+企微+银行合并）
    print('合并资金汇总...')
    fund_rows = _build_fund_summary(wechat_file, alipay_file, bill_file, qiwei_file, bank_file)
    if fund_rows:
        sheets['资金汇总'] = fund_rows
        print(f'  → 资金汇总: {len(fund_rows)}行')

    # 5. 公域账单各渠道汇总 → 原样（含快手/京东）
    _copy_sheets(bill_file, '公域账单清洗.xlsx', {
        '抖音账单汇总': '抖音账单汇总',
        '视频号账单汇总': '视频号账单汇总',
        '百度账单汇总': '百度账单汇总',
        '小红书账单汇总': '小红书账单汇总',
        '快手账单汇总': '快手账单汇总',
        '京东账单汇总': '京东账单汇总',
    }, sheets)

    # 输出
    if sheets:
        output_path = os.path.join(OUTPUT_DIR, '总合并.xlsx')
        write_output(sheets, output_path)
        print(f'\n输出: {output_path}')
        print(f'Sheet数: {len(sheets)}')
    else:
        print('\n无数据可输出（请先运行各模块清洗）')
