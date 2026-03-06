"""数据清洗公共工具函数"""
import os
import sys
import csv
import re
import json
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 缓存文件统一在clean/根目录
if getattr(sys, 'frozen', False):
    _CACHE_DIR = os.path.dirname(sys.executable)
else:
    _CACHE_DIR = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__))))

CACHE_FILE = os.path.join(_CACHE_DIR, '.last_path.json')


def load_cache():
    """读取上次使用的路径"""
    if not os.path.exists(CACHE_FILE):
        return None
    try:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        paths = data.get('paths', [])
        if paths and all(os.path.exists(p) for p in paths):
            return paths
    except:
        pass
    return None


def save_cache(paths):
    """保存本次路径"""
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump({'paths': paths}, f, ensure_ascii=False)


def get_paths_interactive():
    """交互模式获取路径：优先询问是否用缓存"""
    cached = load_cache()
    if cached:
        print(f'\n上次数据源路径:')
        for p in cached:
            print(f'  {p}')
        choice = input('使用上次路径? (Y/回车=是, N=重新输入): ').strip().upper()
        if choice != 'N':
            return cached
    source = input('\n数据源文件夹路径(可拖入): ').strip().strip('"').strip("'")
    if not source:
        print('未输入路径')
        return None
    return [source]


def read_file(filepath, read_config):
    """读取文件为list[dict]，支持excel/csv/auto/加密"""
    fmt = read_config['格式']
    header_row = read_config.get('header行', 0)
    password = None
    if read_config.get('密码') == 'auto':
        password = _extract_password(filepath)

    if fmt == 'auto':
        ext = os.path.splitext(filepath)[1].lower()
        fmt = 'csv' if ext == '.csv' else 'excel'

    if fmt == 'csv':
        if header_row > 0:
            return _read_csv_with_header(filepath, header_row)
        return _read_csv(filepath)
    else:
        return _read_excel(filepath, header_row, password)


def _read_csv(filepath):
    """读取CSV，自动检测编码"""
    for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
        try:
            with open(filepath, 'r', encoding=enc, newline='') as f:
                reader = csv.DictReader(f)
                return [dict(row) for row in reader]
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f'无法解码CSV: {filepath}')


def _read_csv_with_header(filepath, header_row):
    """读CSV但跳过前header_row行，以第header_row行为header"""
    for enc in ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
        try:
            with open(filepath, 'r', encoding=enc, newline='') as f:
                reader = csv.reader(f)
                rows = list(reader)
            if header_row >= len(rows):
                return []
            headers = rows[header_row]
            data = []
            for row in rows[header_row + 1:]:
                d = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
                data.append(d)
            return data
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(f'无法解码CSV: {filepath}')


def _is_ole2(filepath):
    """检查文件是否为OLE2格式（旧版.xls二进制）"""
    try:
        with open(filepath, 'rb') as f:
            return f.read(8) == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
    except:
        return False


def _extract_password(filepath):
    """从文件名提取密码（最后一个空格后的纯字母数字串>=5位）"""
    name = os.path.splitext(os.path.basename(filepath))[0]
    parts = name.rsplit(' ', 1)
    if len(parts) == 2:
        candidate = parts[1].strip()
        if re.match(r'^[A-Za-z0-9]+$', candidate) and len(candidate) >= 5:
            return candidate
    return None


def _read_xls(filepath, header_row=0):
    """用xlrd读取旧版.xls/OLE2格式"""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    if header_row >= ws.nrows:
        return []
    headers = [str(ws.cell_value(header_row, c)) for c in range(ws.ncols)]
    data = []
    for r in range(header_row + 1, ws.nrows):
        d = {}
        for c in range(ws.ncols):
            if c < len(headers) and headers[c]:
                d[headers[c]] = ws.cell_value(r, c)
        data.append(d)
    return data


def _read_excel(filepath, header_row=0, password=None):
    """读取excel，自动判断格式：加密xlsx / 普通xlsx / xls+OLE2 / fallback csv"""
    if password:
        import msoffcrypto
        import io as _io
        with open(filepath, 'rb') as f:
            dec = msoffcrypto.OfficeFile(f)
            dec.load_key(password=password)
            buf = _io.BytesIO()
            dec.decrypt(buf)
            buf.seek(0)
        wb = load_workbook(buf, data_only=True)
    elif _is_ole2(filepath):
        return _read_xls(filepath, header_row)
    else:
        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception:
            return _read_csv_with_header(filepath, header_row)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if header_row >= len(rows):
        return []
    headers = [str(h) if h is not None else '' for h in rows[header_row]]
    data = []
    for row in rows[header_row + 1:]:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = val
        data.append(d)
    return data


def _clean_str(val):
    """清理字符串中的不可见字符+货币符号前缀"""
    if not isinstance(val, str):
        return val
    if val.startswith('¥') or val.startswith('￥'):
        val = val[1:]
    return val.replace('\t', '').replace('\r', '').replace('\n', '').replace('\xa0', ' ').strip()


def clean_strings(data):
    """清理所有字符串字段的不可见字符"""
    for row in data:
        for k in row:
            row[k] = _clean_str(row[k])
    return data


def _not_empty(val):
    if val is None:
        return False
    if isinstance(val, str) and val.strip() in ('', '-'):
        return False
    return True


def remove_empty_rows(data):
    """删除所有字段都为空/None/空字符串的行"""
    return [row for row in data if any(_not_empty(v) for v in row.values())]


def _parse_yyyymm(val):
    """从各种日期格式提取yyyymm"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y%m')
    s = str(val).strip()
    if not s:
        return ''
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d']:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime('%Y%m')
        except ValueError:
            continue
    digits = re.sub(r'\D', '', s)
    if len(digits) >= 6:
        return digits[:6]
    return ''


def _parse_date(val):
    """从各种日期格式提取 (yyyymm, yyyy-mm-dd) 二元组，失败返回 ('', '')"""
    if val is None:
        return '', ''
    if isinstance(val, datetime):
        return val.strftime('%Y%m'), val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return '', ''
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d']:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime('%Y%m'), dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return '', ''


def _normalize_value(val):
    """标准化值用于引用键拼接：统一转float字符串保证精度一致匹配。"""
    if val is None:
        return ''
    s = str(val).replace(',', '').replace('¥', '').replace('￥', '').strip()
    try:
        return str(float(s))
    except (ValueError, TypeError):
        return s


def _to_number(val):
    """将可能带货币符号/逗号的值转为float，失败返回None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('¥', '').replace('￥', '').replace(',', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _auto_column_width(ws):
    """自动调整列宽，中文字符按2宽度计算"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val_str = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def write_output(sheets_data, output_path):
    """将 {sheet名: list[dict]} 写入xlsx，自动调整列宽"""
    fname = os.path.basename(output_path)
    total = sum(1 for rows in sheets_data.values() if rows)
    print(f'\n  写入 {fname} ({total}个sheet)...')
    wb = Workbook()
    first = True
    idx = 0
    for sheet_name, rows in sheets_data.items():
        if not rows:
            continue
        idx += 1
        print(f'    [{idx}/{total}] {sheet_name} ({len(rows)}行)')
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        headers = list(rows[0].keys())
        for j, h in enumerate(headers):
            ws.cell(1, j + 1, h)
        for i, row in enumerate(rows):
            for j, h in enumerate(headers):
                val = row.get(h, '')
                if val is None:
                    val = ''
                elif isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d')
                ws.cell(i + 2, j + 1, val)
        _auto_column_width(ws)
        # 添加自动筛选（所有列）
        if headers:
            from openpyxl.utils import get_column_letter
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f'A1:{last_col}{len(rows) + 1}'
    if first:
        ws = wb.active
        ws.title = '无数据'
    print(f'    保存中...')
    _save_with_retry(wb, output_path)


def _save_with_retry(wb, path, msg=''):
    """保存workbook，遇到文件被占用时提示关闭后重试"""
    while True:
        try:
            wb.save(path)
            return
        except PermissionError:
            fname = os.path.basename(path)
            print(f'\n*** 文件被占用: {fname} ***')
            print(f'请关闭Excel中的 {fname}，然后按回车重试...')
            input()


def read_xml_spreadsheet(filepath):
    """解析XML Spreadsheet 2003格式(.xls实际为XML)，返回list[dict]

    支付宝导出的.xls文件实际是XML Spreadsheet 2003格式，
    非标准BIFF格式，xlrd/openpyxl均无法读取。
    前2行为注释行(#账号、#查询日期)，第3行为列名，后续为数据。
    """
    import xml.etree.ElementTree as ET
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

    tree = ET.parse(filepath)
    root = tree.getroot()

    all_data = []
    for ws in root.findall('.//ss:Worksheet', ns):
        table = ws.find('ss:Table', ns)
        if table is None:
            continue
        rows = table.findall('ss:Row', ns)
        if len(rows) < 3:
            continue

        # 跳过前2行注释，第3行(index=2)为列名
        header_cells = rows[2].findall('ss:Cell', ns)
        headers = []
        for c in header_cells:
            data = c.find('ss:Data', ns)
            headers.append(data.text.strip() if data is not None and data.text else '')

        # 数据行(index=3起)
        for row in rows[3:]:
            cells = row.findall('ss:Cell', ns)
            record = {}
            for i, c in enumerate(cells):
                if i >= len(headers):
                    break
                data = c.find('ss:Data', ns)
                val = data.text.strip() if data is not None and data.text else ''
                record[headers[i]] = val
            for h in headers:
                if h not in record:
                    record[h] = ''
            all_data.append(record)

    return all_data


def unzip_and_read_csvs(zip_path, encoding='gbk'):
    """解压ZIP中的CSV文件并合并为list[dict]，跳过非CSV文件"""
    import zipfile, io as _io
    rows = []
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for name in zf.namelist():
            if not name.lower().endswith('.csv'):
                continue
            with zf.open(name) as f:
                raw = f.read()
                for enc in ([encoding] if encoding else []) + ['utf-8-sig', 'utf-8', 'gbk', 'gb18030']:
                    try:
                        text = raw.decode(enc)
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                else:
                    text = raw.decode('gbk', errors='replace')
                reader = csv.DictReader(_io.StringIO(text))
                for row in reader:
                    rows.append(dict(row))
    return rows


def read_output_sheets(path):
    """读取已有xlsx，返回 {sheet名: list[dict]}"""
    if not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f'  ⚠ 存档文件损坏，将忽略: {os.path.basename(path)} ({e})')
        return {}
    result = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            result[sn] = []
            continue
        headers = [str(h) if h is not None else '' for h in rows[0]]
        data = []
        for row in rows[1:]:
            d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}
            data.append(d)
        result[sn] = data
    wb.close()
    return result
